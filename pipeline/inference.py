#!/usr/bin/env python3
"""
================================================================================
LinAlg-Bench · Stage 1 — Inference with LLM Equivalence Check (inference.py)
================================================================================

Title:       0-shot inference with post-hoc LLM equivalence verification
Version:     2.0 (r2 — with LLM equivalence check)
Date:        2026-04-11

PURPOSE:
    Runs bare 0-shot inference across all linear algebra
    subcategories with novel LLM equivalence check post-processing:
      determinant · eigenvalue · rank · nullity
      multiplication · matrix_power · matrix_vector · transpose

    Key features:
      • System prompt and task text per subcategory (SUBCAT_CONFIG)
      • Parallel API calls with per-provider rate limiting
      • LLM equivalence check (post-inference): re-verifies extraction failures
        using DeepSeek to detect correct answers despite extraction code bugs
      • Adaptive token step-up on truncation (ceiling per token_mult)
      • Loop detection and infinite repetition trimming
      • Result deduplication by (question_id, variant) with preference for correct==True
      • Atomic per-call result saves with timestamp-based deduplication

    Equivalence check: When extraction fails (correct==None or correct==False),
    automatically uses DeepSeek to verify if the model's response actually
    contains the mathematically correct answer despite code extraction failure.

PARALLELIZATION:
    • ThreadPoolExecutor: Concurrent API calls (configurable workers, default 8)
    • Semaphore-based rate limiting: Separate semaphores per provider
    • Heap buffer: Preserves result ordering by (question_id, variant)
    • Thread-safe JSONL: Atomic writes with file locks prevent interleaving
    • Per-call result saves: Each task result written immediately with timestamp

TRANSPORT (inference_llm.py):
    • Backend auto-detection: OpenRouter / OpenAI / GenAI
    • Adaptive token step-up: truncations retry with increased max_tokens
    • Loop detection + trimming: Flags infinite repetition, appends sentinel
    • finish_reason: "stop" | "length" | "loop_trimmed" | None (recorded per call)

USAGE EXAMPLES:
    # Standard: inference on rank, DeepSeek-V3
    python pipeline/inference.py \\
        --input data/sample/sample_3x3.csv \\
        --model DeepSeek-V3 \\
        --output ./results

    # With custom parameters
    python pipeline/inference.py \\
        --input data/sample/sample_3x3.csv \\
        --model GPT-4o \\
        --output ./results \\
        --max-workers 4 \\
        --rate-limit 0.5

    # Resume: continue interrupted run, skip completed (id, variant) pairs
    python pipeline/inference.py \\
        --input data/sample/sample_3x3.csv \\
        --model DeepSeek-V3 \\
        --output ./results \\
        --resume all

    # Resume failures only: redo extraction failures (correct==None)
    python pipeline/inference.py \\
        --input data/sample/sample_3x3.csv \\
        --model DeepSeek-V3 \\
        --output ./results \\
        --resume failures

    # Dry-run: preview prompts without API calls
    python pipeline/inference.py \\
        --input data/sample/sample_3x3.csv \\
        --model DeepSeek-V3 \\
        --output ./results \\
        --dry-run

OUTPUT:
    {output}/{model}_results.jsonl — per-call results (one JSON per line):
        • question_id       : Problem identifier from input
        • variant           : Always "standard"
        • model             : LLM model name
        • response          : Full model response text
        • extracted_answer  : Extracted final answer (or None if extraction failed)
        • correct           : Boolean (extracted == ground_truth, or LLM equivalence result)
        • ground_truth      : Correct answer value
        • answer_latex      : Correct answer in LaTeX (used for equivalence check)
        • problem_latex     : Problem statement in LaTeX
        • timestamp         : ISO 8601 timestamp of API call
        • finish_reason     : "stop" | "length" | "loop_trimmed" | None

    {output}/{model}_summary.csv — per-question accuracy summary:
        • question_id  : Problem identifier
        • total_calls  : Number of API calls for this question
        • correct_count: Count where correct==True
        • accuracy     : correct_count / total_calls (percentage)

EQUIVALENCE CHECK (post-inference, automatic):
    Triggers automatically on records where correct==None or correct==False.
    Uses DeepSeek via OpenRouter to verify if model's response contains the
    mathematically correct answer despite extraction code failure. Updates
    the correct field in JSONL based on LLM verification result.

    Pattern: EQUIVALENCE_PROMPT with strict verification rules for scalars,
    vectors, matrices, and eigenvalues. Conservative: returns FALSE when in doubt.

ENVIRONMENT:
    export OPENROUTER_API_KEY="your-key"   # for OpenRouter models (required)
    export OPENAI_API_KEY="your-key"       # for OpenAI models (if needed)
    export GEMINI_API_KEY="your-key"       # for Gemini models (optional)

================================================================================
"""

import os
import re
import json
import time
import logging
import argparse
import numpy as np
import pandas as pd
from datetime import datetime
from typing import Dict, List, Optional
import heapq
from threading import Lock, Semaphore
from concurrent.futures import ThreadPoolExecutor, as_completed, TimeoutError as FutureTimeoutError
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from inference_llm import InferenceClient
from subcat_config import get_config as get_subcat_config
from models import MODELS, get_model_names
try:
    from json_repair import loads as json_repair_loads
except ImportError:
    json_repair_loads = None

try:
    from tqdm import tqdm
except ImportError:
    def tqdm(x, **kwargs):
        return x


# ═════════════════════════════════════════════════════════════════════════
# 1. MODEL CONFIG
# ═════════════════════════════════════════════════════════════════════════

VARIANTS = [
    "standard",
]


# ═════════════════════════════════════════════════════════════════════════
# 3. MATRIX PARSING & FORMATTING
# ═════════════════════════════════════════════════════════════════════════

def parse_latex_matrix(latex_str: str) -> np.ndarray:
    clean_json_str = re.sub(r'\\(?![u"bfnrt])', r'\\\\', str(latex_str))
    latex_str = str(clean_json_str)
    content = re.search(
        r'\\begin\{bmatrix\}(.+?)\\end\{bmatrix\}', latex_str, re.DOTALL
    )
    if not content:
        raise ValueError(f"Cannot parse LaTeX: {latex_str[:80]}...")
    rows = []
    for r in content.group(1).split('\\\\'):
        r = r.strip()
        if r:
            rows.append([int(x.strip()) for x in r.replace("\\", "").split('&')])
    return np.array(rows, dtype=int)


def matrix_to_latex(A) -> str:
    if isinstance(A, np.ndarray):
        A = A.tolist()
    rows = [" & ".join(str(int(x)) for x in row) for row in A]
    return "\\begin{bmatrix} " + " \\\\ ".join(rows) + " \\end{bmatrix}"


# ═════════════════════════════════════════════════════════════════════════
# 4. SUBCAT CONFIG — system prompts and task text per subcategory
# ═════════════════════════════════════════════════════════════════════════

_BASE_SYSTEM = (
    "You are a precise mathematical assistant. "
    "Show all computation steps clearly. "
)

SUBCAT_CONFIG = {
    # Canonical short forms: det, eig, rank, null, mult, pow, vec, trans, trace
    # Mandatory columns: id, answer_latex, problem_latex, problem_text, subcategory
    # system_prompt: locked instruction set, includes boxed directive
    # boxed_hint: fallback if problem_text is empty/missing
    # token_mult: ceiling = min_tokens × token_mult (adaptive step-up upper bound)
    "det": {
        "system_prompt": _BASE_SYSTEM + "Always put your final numerical answer inside \\boxed{}.",
        "boxed_hint":    "Put your final numerical answer inside \\boxed{{}} at the end.",
        "token_mult":    3,
    },
    "eig": {
        "system_prompt": _BASE_SYSTEM + "Always put your final answers inside \\boxed{}.",
        "boxed_hint":    "Put your final eigenvalue(s) inside \\boxed{{}} at the end.",
        "token_mult":    4,
    },
    "rank": {
        "system_prompt": _BASE_SYSTEM + "Always put your final integer answer inside \\boxed{}.",
        "boxed_hint":    "Put your final integer answer inside \\boxed{{}} at the end.",
        "token_mult":    2,
    },
    "nullity": {
        "system_prompt": _BASE_SYSTEM + "Always put your final integer answer inside \\boxed{}.",
        "boxed_hint":    "Put your final integer answer inside \\boxed{{}} at the end.",
        "token_mult":    2,
    },
    "mult": {
        "system_prompt": _BASE_SYSTEM + "Always put your final matrix answer inside \\boxed{}.",
        "boxed_hint":    "Put your final matrix answer inside \\boxed{{}} at the end.",
        "token_mult":    2,
    },
    "pow": {
        "system_prompt": _BASE_SYSTEM + "Always put your final matrix answer inside \\boxed{}.",
        "boxed_hint":    "Put your final matrix answer inside \\boxed{{}} at the end.",
        "token_mult":    2,
    },
    "vec": {
        "system_prompt": _BASE_SYSTEM + "Always put your final vector answer inside \\boxed{}.",
        "boxed_hint":    "Put your final vector answer inside \\boxed{{}} at the end.",
        "token_mult":    2,
    },
    "trans": {
        "system_prompt": _BASE_SYSTEM + "Always put your final matrix answer inside \\boxed{}.",
        "boxed_hint":    "Put your final matrix answer inside \\boxed{{}} at the end.",
        "token_mult":    2,
    },
    "trace": {
        "system_prompt": _BASE_SYSTEM + "Always put your final numerical answer inside \\boxed{}.",
        "boxed_hint":    "Put your final numerical answer inside \\boxed{{}} at the end.",
        "token_mult":    2,
    },
}


# ═════════════════════════════════════════════════════════════════════════
# 5. PROMPT BUILDER
# ═════════════════════════════════════════════════════════════════════════

def build_standard_prompt(problem_text: str, problem_latex: str, subcat: str) -> str:
    """
    Bare 0-shot prompt built from Excel columns.
    problem_text: task instruction already includes the problem description
    problem_latex: matrix/vector in LaTeX form
    subcat: subcategory name (can be short form: det, eigen, pow, transp, vector, nullity, rank)
    """
    # Normalize subcat (det → determinant, etc.)
    subcat_full = _normalize_subcat(subcat)
    cfg = SUBCAT_CONFIG.get(subcat_full)
    if not cfg:
        raise ValueError(f"Unknown subcategory: {subcat} (normalized: {subcat_full})")

    # If problem_text is missing/NaN, fall back to boxed_hint
    _pt = str(problem_text).strip() if problem_text is not None else ""
    task_part = _pt if _pt and _pt.lower() not in ("nan", "none") else cfg['boxed_hint']
    return (
        f"Solve this linear algebra problem. Show your work and give the final answer.\n\n"
        f"{task_part}\n\n"
        f"{problem_latex}\n\n"
    )


# ═════════════════════════════════════════════════════════════════════════
# 6. ANSWER EXTRACTION
# ═════════════════════════════════════════════════════════════════════════
# ground_truth_fn and extract_answer_fn are provided per-subcat by
# subcat_config.get_config(subcat) — no duplication here.

def count_boxed(text: str) -> int:
    """Count \\boxed{} occurrences — multiples indicate hedging behavior."""
    if not text:
        return 0
    return len(re.findall(r'\\boxed\{', text))


# ═════════════════════════════════════════════════════════════════════════
# 6b. LLM EQUIVALENCE CHECK — post-inference fallback for extraction failures
# ═════════════════════════════════════════════════════════════════════════

EQUIVALENCE_PROMPT = """You are a precise mathematics verifier. Compare the correct answer with the model's final answer and determine if they are mathematically equivalent.

CORRECT ANSWER (ground truth):
{answer_latex}

MODEL'S RESPONSE (last 10 lines):
---
{last_lines}
---

VERIFICATION RULES (STRICT):

1. SCALARS (rank, nullity, trace):
   - Match integers: 5 == 5, 0 == 0
   - Match decimals: -8.9305 == -8.9305
   - Match equivalent forms: 5 == 5.0, -3/4 == -0.75, -0.75 == -75/100
   - Match rounded vs exact: if correct is 4.3, model's 4 or 4.3 is WRONG
   - Model must provide explicit numeric answer (prose like "rank is 5" counts)

2. VECTORS (matrix_vector):
   - Same dimension required
   - Each element must match (within rounding tolerance)
   - Format variants OK: [1,2,3], (1,2,3), column vector, etc.
   - Must be complete — missing even one element = WRONG

3. MATRICES (multiplication, matrix_power, transpose):
   - Same dimensions required (m×n)
   - Each element must match (within tolerance)
   - Format variants OK: LaTeX matrix, [row1; row2], etc.
   - Full matrix required — partial matrix = WRONG
   - Matrix structure must be intact (5 rows × 5 cols for 5×5)

4. EIGENVALUES (eigen):
   - Same set of eigenvalues required
   - Order doesn't matter (e.g., {{-2.0, 3.0, 5.0}} equivalent to {{5.0, -2.0, 3.0}})
   - All values must be present
   - Accept decimal/radical equivalence: 4+2√5 ≈ 8.47
   - If model provides integers instead of floats (e.g., -7,-3,2,6,8 for correct -8.93,-4.38,-2.01,9.73,11.59) = WRONG

5. ACCEPTABLE EQUIVALENCES:
   - Simplification: -6/8 == -3/4
   - Decimal forms: 0.5 == 1/2
   - Negative zero: -0.0 == 0.0
   - Scientific notation: 1e-5 == 0.00001
   - Rounding (±1 in last decimal place): 3.14159... ≈ 3.14 or 3.15 (BORDERLINE — err on WRONG)

6. MISMATCH INDICATORS (return false):
   - Answer partially provided but incomplete
   - Wrong dimension (2×3 when should be 5×5)
   - Identity matrix when should be actual computation
   - Prose answer without numeric value (e.g., "full rank" without saying "5")
   - Different eigenvalue count
   - Sign error (5 vs -5)
   - Clearly stated wrong value ("the rank is 4" when correct is 5)

DECISION LOGIC:
- TRUE: Model's answer is mathematically equivalent to ground truth
- FALSE: Model's answer is wrong, incomplete, or not found
- When in doubt, return FALSE (conservative)

Respond with ONLY this JSON (no markdown):
{{"equivalent": true, "reason": "specific evidence: e.g., all 5 eigenvalues match, matrix dimensions 5×5 with correct elements"}}
or
{{"equivalent": false, "reason": "specific reason: e.g., eigenvalues are integers not floats, missing nullity computation, wrong matrix dimensions"}}"""


def check_equivalence_llm(answer_latex: str, response: str, api_key: str = None) -> tuple:
    """
    Use DeepSeek via OpenRouter to check if model's final answer matches answer_latex.
    Uses last 10 lines of response. Returns (is_equivalent: bool, reason: str).
    """
    if api_key is None:
        api_key = os.environ.get("OPENROUTER_API_KEY", "")
    if not api_key:
        return False, "No OPENROUTER_API_KEY"

    # Last 10 lines
    lines = response.splitlines() if response else []
    last_lines = "\n".join(lines[-10:]) if lines else "(empty)"

    # Escape curly braces in values so they're not interpreted as format placeholders
    user_prompt = EQUIVALENCE_PROMPT.format(
        answer_latex=answer_latex.replace('{', '{{').replace('}', '}}'),
        last_lines=last_lines.replace('{', '{{').replace('}', '}}'),
    )

    try:
        from openai import OpenAI
        client = OpenAI(base_url="https://openrouter.ai/api/v1", api_key=api_key)

        resp = client.chat.completions.create(
            model="deepseek/deepseek-v3.2",
            messages=[{"role": "user", "content": user_prompt}],
            temperature=0.0,
            max_tokens=200,
        )

        text = resp.choices[0].message.content or ""
        text = re.sub(r'^```(?:json)?\s*', '', text.strip(), flags=re.IGNORECASE)
        text = re.sub(r'\s*```$', '', text)

        loader = json_repair_loads if json_repair_loads else json.loads
        result = loader(text)
        return bool(result.get("equivalent", False)), str(result.get("reason", ""))

    except Exception as e:
        return False, f"LLM check failed: {e}"


def rerun_equivalence_check(results_file: str, api_key: str = None, dry_run: bool = False):
    """
    Post-inference: re-check records where:
      1. correct == None (extraction returned None)
      2. correct == False (extraction may have grabbed wrong value, verify if correct answer is in response)

    Uses answer_latex + last 10 lines of response. Updates JSONL in-place.
    """
    if not os.path.exists(results_file):
        print(f"  No results file: {results_file}")
        return

    records = []
    with open(results_file) as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    records.append(json.loads(line))
                except json.JSONDecodeError:
                    pass

    if not records:
        print("  No records to re-check")
        return

    # Check both: correct is None (extraction failed) OR correct is False (extraction may have grabbed wrong value)
    check_records = [r for r in records if r.get("correct") is None or r.get("correct") is False]
    if not check_records:
        print(f"  No records need equivalence check (all {len(records)} records have correct=True)")
        return

    print(f"  Re-checking {len(check_records)} extraction failures/anomalies with LLM equivalence...")

    corrected_count = 0
    for r in check_records:
        qid = r.get("question_id", "?")
        answer_latex = r.get("answer_latex", "")
        resp = r.get("response", "")
        extracted = r.get("extracted_answer")
        correct_before = r.get("correct")

        if dry_run:
            print(f"  [DRY RUN] {qid}: correct={correct_before}, extracted={extracted}")
            continue

        is_eq, reason = check_equivalence_llm(answer_latex, resp, api_key)
        if is_eq:
            print(f"  ✓ {qid}: EQUIVALENT (was {correct_before}) — {reason}")
            r["correct"] = True
            r["equivalence_check"] = "LLM_CONFIRMED"
            r["equivalence_reason"] = reason
            if correct_before is False:
                corrected_count += 1
        else:
            print(f"  ✗ {qid}: NOT EQUIVALENT — {reason}")
            if "equivalence_check" not in r:
                r["equivalence_check"] = "LLM_REJECTED"
            r["equivalence_reason"] = reason

    with open(results_file, "w") as f:
        for r in records:
            f.write(json.dumps(r, default=str) + "\n")

    confirmed = sum(1 for r in check_records if r.get("correct") is True)
    print(f"  Done: {confirmed}/{len(check_records)} confirmed as equivalent (corrected {corrected_count} from False→True)")


# Loop detection, loop trimming, adaptive tokens, finish_reason:
# all handled inside InferenceClient (inference_llm.py)


# ═════════════════════════════════════════════════════════════════════════
# 7. DATA LOADING
# ═════════════════════════════════════════════════════════════════════════

def load_judge_labels(judge_path: str) -> Dict[str, Dict]:
    """
    Load forensic judge error classifications.
    Expected columns: Problem_ID, Error_Tag, Sign_Error_Types, Root_Cause_Hypothesis
    """
    if not judge_path or not os.path.exists(judge_path):
        logging.warning("No judge labels file — results will lack error type breakdown")
        return {}

    df = pd.read_excel(judge_path) if judge_path.endswith('.xlsx') else pd.read_csv(judge_path)

    labels = {}
    for _, row in df.iterrows():
        qid = str(row.get("Problem_ID", row.get("id", ""))).strip()
        labels[qid] = {
            "error_tag": row.get("Error_Tag", "UNKNOWN"),
            "sign_subtype": row.get("Sign_Error_Types", ""),
            "root_cause": row.get("Root_Cause_Hypothesis", ""),
        }
    logging.info(f"Loaded judge labels for {len(labels)} questions")
    return labels


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Map common column name variants to standard internal names."""
    df = df.copy()
    df.columns = df.columns.str.lower().str.strip()

    mapping = {
        'id': ['question_id', 'problem_id', 'problemid', 'qid', 'idx'],
        'problem_text': ['question', 'problem', 'text', 'prompt', 'task', 'instruction'],
        'problem_latex': ['latex', 'matrix', 'equation', 'problem_ascii', 'ascii'],
        'answer_latex': ['answer', 'answer_value', 'solution', 'gt', 'ground_truth', 'ans'],
        'subcategory': ['subcat', 'category', 'task', 'type', 'subcategory'],
    }

    for std_name, aliases in mapping.items():
        if std_name in df.columns:
            continue
        for alias in aliases:
            if alias in df.columns:
                df.rename(columns={alias: std_name}, inplace=True)
                break

    return df


def _detect_file_type(path: str) -> str:
    ext = os.path.splitext(path)[1].lower()
    if ext in ('.xlsx', '.xls'):
        return 'excel'
    elif ext == '.csv':
        return 'csv'
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def load_dataset(input_path: str, subcat_cfg: dict, gt_fn) -> pd.DataFrame:
    """
    Load dataset from CSV or Excel with flexible column name detection.

    Standard internal names:
      id, problem_text, problem_latex, answer_latex, subcategory
    """
    file_type = _detect_file_type(input_path)

    if file_type == 'csv':
        df = pd.read_csv(input_path)
    else:
        df = pd.read_excel(input_path)

    df = _normalize_columns(df)

    # Warn about missing columns but don't fail
    expected = ['id', 'problem_text', 'problem_latex', 'answer_latex']
    for col in expected:
        if col not in df.columns:
            print(f"  [WARN] Column '{col}' not found in input — inference may fail")

    # Subcategory fallback: subcategory → (use CLI arg)
    if 'subcategory' not in df.columns:
        pass  # will be set by CLI arg in run_pipeline

    # Normalize subcategory values
    if 'subcategory' in df.columns:
        df['subcategory'] = df['subcategory'].apply(
            lambda x: _normalize_subcat(str(x).lower().strip()) if pd.notna(x) else x
        )

    gts = []
    for _, row in df.iterrows():
        # Determine ground truth per-row from the row's own subcategory.
        # The passed gt_fn is only used as fallback when no subcategory column exists.
        row_subcat_raw = row.get('subcategory')

        # Check for NaN/None before converting to string
        if pd.notna(row_subcat_raw) and str(row_subcat_raw).strip().lower() not in ('', 'nan', 'none'):
            row_subcat = str(row_subcat_raw).lower().strip()
            sc = get_subcat_config(row_subcat)
            gt = sc.ground_truth_fn(row['answer_latex'])
        elif gt_fn:
            # Fallback: no subcategory column, use passed gt_fn
            gt = gt_fn(row['answer_latex'])
        else:
            raise ValueError(f"No subcategory for row id={row.get('id', '?')} and no gt_fn provided")
        gts.append(gt)
    df["ground_truth"] = gts
    return df


def _normalize_subcat(input_form: str) -> str:
    """
    Normalize any subcategory name to canonical form.
    Maps all variants to: det, eig, rank, nullity, mult, pow, vec, trans, trace
    """
    canonical_map = {
        # Determinant variants
        "det": "det",
        "determinant": "det",

        # Eigenvalue variants
        "eig": "eig",
        "eigen": "eig",
        "eigenvalue": "eig",
        "eigenvalues": "eig",

        # Rank variants
        "rank": "rank",

        # Nullity variants (use complete form — avoids "null" keyword issues)
        "nullity": "nullity",
        "null": "nullity",

        # Multiplication variants
        "mult": "mult",
        "multiplication": "mult",

        # Matrix power variants
        "pow": "pow",
        "pow2": "pow",
        "power": "pow",
        "matrix_power": "pow",

        # Matrix-vector variants
        "vec": "vec",
        "vector": "vec",
        "matrix_vector": "vec",
        "matvec": "vec",

        # Transpose variants
        "trans": "trans",
        "transpose": "trans",
        "transp": "trans",

        # Trace (no variants expected)
        "trace": "trace",
    }
    lower_form = str(input_form).lower().strip()
    return canonical_map.get(lower_form, lower_form)


def load_completed(results_file: str, resume_mode: str = "all") -> set:
    """
    Load set of completed (qid, variant) tuples.

    Uses last-record-wins: if the JSONL has multiple records for the same
    (qid, variant) pair (from retries or interrupted runs), the LAST record's
    state determines whether the task is considered complete.

    resume_mode:
      "all"      → skip all existing records (default --resume)
      "failures" → skip only records with correct!=None, redo correct==None
    """
    if not os.path.exists(results_file):
        return set()

    last_record = {}
    with open(results_file) as f:
        for line in f:
            try:
                r = json.loads(line)
                key = (r["question_id"], r["variant"])
                last_record[key] = r
            except (json.JSONDecodeError, KeyError):
                pass

    if resume_mode == "all":
        return set(last_record.keys())

    completed = set()
    for key, r in last_record.items():
        if r.get("correct") is not None:
            completed.add(key)
    return completed


def save_result(result: Dict, results_file: str, file_lock: Lock = None):
    """Save result to single _results.jsonl — all outcomes including None."""
    if file_lock:
        with file_lock:
            _save_result_unsafe(result, results_file)
    else:
        _save_result_unsafe(result, results_file)


def _save_result_unsafe(result: Dict, results_file: str):
    with open(results_file, "a") as f:
        f.write(json.dumps(result, default=str) + "\n")


def deduplicate_results(results_file: str):
    """
    Deduplicate results by (question_id, variant).
    Strategy:
      1. Prefer records with correct==True
      2. If no True record, keep latest by timestamp
      3. Remove all other responses for the same (id, variant) pair

    Uses pure Python (no pandas) to avoid NaN corruption of None values.
    Overwrites JSONL file with deduplicated records.
    """
    if not os.path.exists(results_file):
        return

    # Load all results
    all_results = []
    with open(results_file) as f:
        for line in f:
            try:
                all_results.append(json.loads(line))
            except json.JSONDecodeError:
                pass

    if not all_results:
        return

    # Group by (question_id, variant)
    groups = {}
    for r in all_results:
        key = (r["question_id"], r["variant"])
        groups.setdefault(key, []).append(r)

    # Pick best record per group
    deduped = []
    for key, records in groups.items():
        # Sort: correct first (True > False > None), then latest timestamp
        def sort_key(r):
            correct_rank = {True: 0, False: 1, None: 2}.get(r.get("correct"), 2)
            ts = r.get("timestamp", "")
            return (correct_rank, ts)
        records.sort(key=sort_key)
        deduped.append(records[0])

    # Sort by (question_id, variant) for consistent ordering
    deduped.sort(key=lambda r: (r["question_id"], r["variant"]))

    # Write back
    with open(results_file, "w") as f:
        for r in deduped:
            f.write(json.dumps(r, default=str) + "\n")

    removed = len(all_results) - len(deduped)
    if removed > 0:
        logging.info(f"  Deduplicated {len(all_results)} → {len(deduped)} records (removed {removed} duplicates, preferred correct==True)")


# ═════════════════════════════════════════════════════════════════════════
# 7b. PARALLEL EXECUTION
# ═════════════════════════════════════════════════════════════════════════

class ParallelExecutor:
    """Execute InferenceClient calls in parallel while maintaining result ordering.

    Features:
      • Heap-buffered flush: results written to JSONL in (task_idx) order
      • Per-provider rate limiting: separate semaphores for OpenRouter, OpenAI, GenAI
      • Task timeout: individual tasks killed after `task_timeout` seconds
    """

    # Per-provider concurrency limits (conservative to avoid 429s)
    PROVIDER_LIMITS = {
        "openrouter": 5,
        "openai":     3,
        "genai":      4,
        "default":    4,
    }

    def __init__(self, client: InferenceClient, results_file: str,
                 extract_fn,
                 max_workers: int = 8, rate_limit: float = 1.0,
                 max_tokens: int = 8192, ceiling: int = 16384,
                 task_timeout: int = 600):
        self.client       = client
        self.results_file = results_file
        self.extract_fn   = extract_fn
        self.max_workers  = max_workers
        self.rate_limit   = rate_limit
        self.max_tokens   = max_tokens
        self.ceiling      = ceiling
        self.task_timeout = task_timeout
        self.call_count   = 0
        self.call_lock    = Lock()
        self.file_lock    = Lock()
        self.buffer_lock  = Lock()
        self.result_buffer = []   # Min-heap: (task_idx, result)
        self.next_flush_idx  = 0

        # Per-provider rate-limiting semaphores
        provider = self._get_provider(client.model_id)
        provider_limit = self.PROVIDER_LIMITS.get(provider, self.PROVIDER_LIMITS["default"])
        effective_workers = min(max_workers, provider_limit)
        self.rate_semaphore = Semaphore(effective_workers)
        logging.info(f"  Rate limit: {provider} (semaphore={effective_workers}, provider_limit={provider_limit})")

    @staticmethod
    def _get_provider(model_id: str) -> str:
        """Determine provider from model ID for rate limiting."""
        mid = model_id.lower()
        if "gemini" in mid:
            return "genai"
        elif "gpt" in mid or "o1" in mid or "o3" in mid:
            return "openai"
        else:
            return "openrouter"

    def _flush_buffer(self):
        """Flush contiguous results to disk in task_idx order.
        NOTE: Caller must hold buffer_lock — this method does not acquire it
        to avoid deadlock when called from within 'with self.buffer_lock'.

        On resume, task_idx values start from 0 again (not contiguous with
        previous run). We flush any record whose task_idx is strictly less
        than the smallest unflushed index in the buffer, ensuring results
        are written in order without gaps blocking the flush.
        """
        while self.result_buffer:
            idx = self.result_buffer[0][0]
            # Flush if this is the next expected index, OR if we've already
            # passed it (resume gap: old run had indices 0..49, new run has
            # 0..19 — we flush immediately since no older indices are pending)
            if idx == self.next_flush_idx or idx < self.next_flush_idx:
                _, result = heapq.heappop(self.result_buffer)
                save_result(result, self.results_file, self.file_lock)
                # Only advance next_flush_idx for indices >= current position
                if idx >= self.next_flush_idx:
                    self.next_flush_idx = idx + 1
            else:
                break  # gap — wait for missing index

    def _execute_task(self, task_idx: int, qid: str,
                     prompt: str, gt: str, problem_latex: str,
                     answer_latex: str, subcat: str):
        try:
            # Build prompt
            full_prompt = prompt

            # Infer subcat if not provided
            if not subcat:
                subcat = _infer_subcat(problem_latex)

            # Get config
            sc        = get_subcat_config(subcat)
            extract_fn = sc.extract_answer_fn

            logging.info(f"  Task {task_idx}: {qid}")

            # Call LLM
            cr = self.client.call(
                user_prompt=full_prompt,
                qid=qid,
                max_tokens=self.max_tokens,
                ceiling=self.ceiling,
            )

            with self.call_lock:
                self.call_count += 1

            # Use per-row extract_fn if subcat provided, else use default
            if subcat:
                sc = get_subcat_config(subcat)
                extract_fn = sc.extract_answer_fn
            else:
                extract_fn = self.extract_fn

            extracted = extract_fn(cr.text) if cr.text else None
            result = {
                "question_id":          qid,
                "variant":              "standard",
                "model":                self.client.model_id,
                "prompt_length":        len(prompt),
                "prompt_tokens_approx": len(prompt) // 4,
                "response":             cr.text,
                "extracted_answer":     extracted,
                "boxed_count":          count_boxed(cr.text) if cr.text else 0,
                "correct":              (extracted == gt) if extracted is not None else None,
                "ground_truth":         gt,
                "problem_latex":        problem_latex,
                "answer_latex":         answer_latex,
                "latency_ms":           cr.latency_ms,
                "tokens_used":          cr.tokens_used,
                "finish_reason":        cr.finish_reason,
                "error":                cr.error,
                "timestamp":            datetime.now().isoformat(),
            }

            # Buffer result for ordered flush
            with self.buffer_lock:
                heapq.heappush(self.result_buffer, (task_idx, result))
                self._flush_buffer()

            print(f"✓ Completed {qid}", flush=True)

        except Exception as e:
            print(f"✗ Task failed {qid}: {e}", flush=True)
            raise

    def run(self, tasks_list: List[tuple], pbar_callback=None):
        """Execute all tasks in parallel, preserving result order via heap buffer.

        Task timeout: individual tasks are given `task_timeout` seconds to complete.
        If exceeded, a TIMEOUT is logged and the task is skipped.
        NOTE: Python threads cannot be forcibly killed — the timed-out thread
        continues running in the background and may still write its result to
        JSONL when it eventually completes. The deduplicate_results() call at
        the end of run_pipeline cleans up any such late arrivals.
        """
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {}
            for task in tasks_list:
                # task = (task_idx, qid, prompt, gt, problem_latex, answer_latex, row_subcat)
                task_idx, qid, prompt, gt, problem_latex, answer_latex, row_subcat = task
                future = executor.submit(self._execute_task, task_idx, qid, prompt, gt, problem_latex, answer_latex, row_subcat)
                futures[future] = (task_idx, qid)

            for future in as_completed(futures):
                task_idx, qid = futures[future]
                try:
                    future.result(timeout=self.task_timeout)
                except FutureTimeoutError:
                    logging.error(f"⏱ TIMEOUT {qid} after {self.task_timeout}s")
                except Exception as e:
                    logging.error(f"Failed {qid}: {e}")
                if pbar_callback:
                    pbar_callback(1)

        # Final flush for any remaining buffered results
        time.sleep(0.5)
        with self.buffer_lock:
            while self.result_buffer:
                idx, result = heapq.heappop(self.result_buffer)
                save_result(result, self.results_file, self.file_lock)
                self.next_flush_idx += 1


# ═════════════════════════════════════════════════════════════════════════
# 8. QUICK RESULTS SUMMARY (printed to stdout immediately after API calls)
# ═════════════════════════════════════════════════════════════════════════

def print_results_summary(results_file: str):
    """Print a quick summary of results to stdout after API calls complete."""
    if not os.path.exists(results_file):
        return

    all_results = []
    try:
        with open(results_file) as f:
            for line in f:
                try:
                    all_results.append(json.loads(line))
                except json.JSONDecodeError:
                    pass
    except Exception:
        return

    if not all_results:
        return

    # Count by correctness and subcategory
    correct_count = 0
    subcats = {}

    for record in all_results:
        qid = record.get('question_id', '')
        correct = record.get('correct', None)

        # Extract subcat from qid (e.g., "ds4-C_5x5_det_001" → "det")
        parts = qid.split('_')
        subcat = parts[-2] if len(parts) >= 2 else "unknown"

        if subcat not in subcats:
            subcats[subcat] = {"correct": 0, "total": 0}

        subcats[subcat]["total"] += 1
        if correct is True:
            correct_count += 1
            subcats[subcat]["correct"] += 1

    total = len(all_results)
    pct = 100 * correct_count / total if total > 0 else 0

    print(f"\n{'='*60}")
    print(f"  RESULTS SUMMARY")
    print(f"{'='*60}")
    print(f"  Total records: {total}")
    print(f"  Correct: {correct_count}/{total} ({pct:.1f}%)")
    print(f"\n  By subcategory:")
    for subcat in sorted(subcats.keys()):
        stats = subcats[subcat]
        sub_pct = 100 * stats["correct"] / stats["total"] if stats["total"] > 0 else 0
        print(f"    {subcat:10} {stats['correct']}/{stats['total']:2}  ({sub_pct:5.1f}%)")
    print(f"{'='*60}\n")


# ═════════════════════════════════════════════════════════════════════════
# 9. ANALYSIS
# ═════════════════════════════════════════════════════════════════════════

def run_analysis(results_file: str, judge: Dict, model_name: str, output_dir: str):
    """Load all results and produce recovery table by variant and error type."""

    all_results = []
    with open(results_file) as f:
        for line in f:
            try:
                all_results.append(json.loads(line))
            except json.JSONDecodeError:
                pass

    # Build per-question summary
    questions = {}
    for r in all_results:
        qid = r["question_id"]
        if qid not in questions:
            questions[qid] = {"id": qid, "ground_truth": r.get("ground_truth")}
        v = r["variant"]
        # Preserve None vs False distinction — None means extraction failure, not wrong
        raw_correct = r.get("correct")
        questions[qid][f"{v}_correct"] = bool(raw_correct) if raw_correct is not None else None
        questions[qid][f"{v}_answer"] = r.get("extracted_answer")
        questions[qid][f"{v}_boxed_count"] = r.get("boxed_count", 0)
        questions[qid][f"{v}_tokens"] = r.get("tokens_used", 0)

    for qid, q in questions.items():
        lbl = judge.get(qid, {})
        q["error_tag"] = lbl.get("error_tag", "UNKNOWN")
        q["sign_subtype"] = lbl.get("sign_subtype", "")
        q["root_cause"] = lbl.get("root_cause", "")

    summary_df = pd.DataFrame(questions.values())

    # ── Accuracy by variant ──────────────────────────────────────
    logging.info("\n" + "═" * 60)
    logging.info("  DIAGNOSTIC LADDER — ACCURACY BY VARIANT")
    logging.info("═" * 60)
    logging.info(f"  {'Variant':<25} {'Correct':>7} {'Total':>7} {'Accuracy':>10}")
    logging.info("  " + "─" * 55)

    for v in VARIANTS:
        col = f"{v}_correct"
        if col not in summary_df.columns:
            continue
        subset = summary_df[summary_df[col].notna()]
        correct = int(subset[col].fillna(False).astype(bool).sum())
        total = len(subset)
        acc = correct / total * 100 if total else 0
        logging.info(f"  {v:<25} {correct:>7} {total:>7} {acc:>9.1f}%")

    logging.info("═" * 60)

    # ── Recovery by error type ────────────────
    ref_col = "standard_correct"
    if judge and ref_col in summary_df.columns:
        fails_df = summary_df[summary_df[ref_col] != True].copy()  # includes False and None

        logging.info(f"\n  RECOVERY RATES (on {len(fails_df)} failures, by error type)")
        logging.info("  " + "─" * 75)
        header = f"  {'Error Type':<20}"
        for v in VARIANTS:
            header += f"  {v.split('_')[0]:>8}"
        header += f"  {'Count':>6}"
        logging.info(header)
        logging.info("  " + "─" * 75)

        for tag in sorted(fails_df["error_tag"].unique()):
            subset = fails_df[fails_df["error_tag"] == tag]
            n = len(subset)
            row_str = f"  {tag:<20}"
            for v in VARIANTS[:-1]:
                col = f"{v}_correct"
                rec = int(subset[col].sum()) if col in subset.columns else 0
                pct = rec / n * 100 if n else 0
                row_str += f"  {pct:>7.1f}%"
            row_str += f"  {n:>6}"
            logging.info(row_str)

        logging.info("  " + "─" * 75)

        # Sign error subtype drill-down
        sign_fails = fails_df[fails_df["error_tag"] == "SIGN_ERROR"]
        if len(sign_fails) > 0 and sign_fails["sign_subtype"].notna().any():
            logging.info(f"\n  SIGN ERROR SUBTYPE DRILL-DOWN ({len(sign_fails)} cases)")
            logging.info("  " + "─" * 75)

            for st in sorted(sign_fails["sign_subtype"].dropna().unique()):
                if not st:
                    continue
                subset = sign_fails[sign_fails["sign_subtype"] == st]
                n = len(subset)
                row_str = f"  {st:<20}"
                for v in VARIANTS[:-1]:
                    col = f"{v}_correct"
                    rec = int(subset[col].sum()) if col in subset.columns else 0
                    pct = rec / n * 100 if n else 0
                    row_str += f"  {pct:>7.1f}%"
                row_str += f"  {n:>6}"
                logging.info(row_str)

            logging.info("  " + "─" * 75)

    # ── Accuracy note ────────────────────────────────────────────
    logging.info("\n  No scaffolding required.\n")

    # ── Save summary CSV ─────────────────────────────────────────
    summary_csv = os.path.join(output_dir, f"{model_name}_summary.csv")
    summary_df.to_csv(summary_csv, index=False)
    logging.info(f"\n  Saved: {summary_csv}")

    return summary_df


# ═════════════════════════════════════════════════════════════════════════
# 10. MAIN PIPELINE
# ═════════════════════════════════════════════════════════════════════════

def run_pipeline(args):

    # ── Set seed for reproducibility ─────────────────────────────
    if args.seed is not None:
        import random
        random.seed(args.seed)
        import numpy as np
        np.random.seed(args.seed)

    # ── Load dataset first to detect subcategory ─────────────────
    file_type = _detect_file_type(args.input)
    if file_type == 'csv':
        df_raw = pd.read_csv(args.input)
    else:
        df_raw = pd.read_excel(args.input)
    df_raw = _normalize_columns(df_raw)

    # Detect subcat from dataset (subcategory → category → type)
    if 'subcategory' in df_raw.columns:
        subcat_vals = df_raw['subcategory'].dropna().unique()
    elif 'category' in df_raw.columns:
        subcat_vals = df_raw['category'].dropna().unique()
    elif 'type' in df_raw.columns:
        subcat_vals = df_raw['type'].dropna().unique()
    else:
        raise ValueError("Dataset must have 'subcategory', 'category', or 'type' column")

    if len(subcat_vals) == 0:
        raise ValueError("No subcategory found in dataset")

    normalized_vals = [_normalize_subcat(str(v).lower().strip()) for v in subcat_vals]
    unique_subcats  = list(dict.fromkeys(normalized_vals))   # preserve order, dedupe

    if len(unique_subcats) > 1:
        # Multi-subcat dataset: use input filename stem as the output label
        input_stem = os.path.splitext(os.path.basename(args.input))[0]
        subcat = input_stem
        logging.info(f"  Detected subcategory: multi ({', '.join(unique_subcats)}) → label '{subcat}'")
    else:
        subcat = unique_subcats[0]
        logging.info(f"  Detected subcategory: {subcat}")

    # ── Subcat config ────────────────────────────────────────────
    if subcat in SUBCAT_CONFIG:
        subcat_cfg    = SUBCAT_CONFIG[subcat]
        system_prompt = subcat_cfg["system_prompt"]
        sc            = get_subcat_config(subcat)
        gt_fn         = sc.ground_truth_fn
        extract_fn    = sc.extract_answer_fn
    else:
        # Multi-subcat: generic config — per-row subcat handles GT + extraction
        subcat_cfg    = {"token_mult": 3, "boxed_hint": "Put your final answer inside \\boxed{{}}."}
        system_prompt = _BASE_SYSTEM + "Always put your final answer inside \\boxed{}."
        gt_fn         = None
        extract_fn    = None

    # ── Model config ─────────────────────────────────────────────
    model_cfg = MODELS[args.model].copy()
    if args.api_key_env:
        model_cfg["api_key_env"] = args.api_key_env
    if args.api_base:
        model_cfg["api_base"] = args.api_base

    min_tokens  = model_cfg.get("min_tokens", 8192)
    max_tokens  = args.max_tokens if args.max_tokens else min_tokens
    token_mult  = subcat_cfg.get("token_mult", 2)
    ceiling     = max_tokens * token_mult

    # ── Output paths ─────────────────────────────────────────────
    output_dir   = args.output or f"data/{subcat}"
    if args.dry_run:
        output_dir = output_dir + "_dryrun"
    os.makedirs(output_dir, exist_ok=True)
    results_file = os.path.join(output_dir, f"{args.model}_results.jsonl")

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(message)s",
        handlers=[
            logging.FileHandler(os.path.join(output_dir, "inference.log")),
            logging.StreamHandler(),
        ],
    )

    logging.info("=" * 60)
    logging.info(f"  LinAlg-Bench · Unified Inference")
    logging.info(f"  Subcat: {subcat}")
    logging.info(f"  Model:  {args.model}")
    logging.info(f"  Mode:   {'DRY RUN' if args.dry_run else 'LIVE'}")
    logging.info(f"  Tokens: {max_tokens} × {token_mult} = ceiling {ceiling}")
    logging.info("=" * 60)

    # ── API key validation (fail fast with clear message) ────────
    if not args.dry_run:
        api_key_env = model_cfg.get("api_key_env", "")
        api_key = os.environ.get(api_key_env, "") if api_key_env else ""
        if not api_key:
            raise ValueError(
                f"API key is empty for model '{args.model}'!\n"
                f"  Environment variable: {api_key_env}\n"
                f"  Set it with: export {api_key_env}='your-key'\n"
                f"  Or create a .env file with {api_key_env}=your-key"
            )

    # ── Build InferenceClient ────────────────────────────────────
    client = InferenceClient(
        model_cfg=model_cfg,
        system_prompt=system_prompt,
        dry_run=args.dry_run,
    )

    # ── Load data ────────────────────────────────────────────────
    df        = load_dataset(args.input, subcat_cfg, gt_fn)
    judge     = load_judge_labels(args.judge)

    # Pre-resume dedup: clean stale duplicates BEFORE reading completed set
    # so load_completed sees the authoritative last-record state
    if args.resume and results_file and os.path.exists(results_file):
        deduplicate_results(results_file)

    completed = load_completed(results_file, args.resume) if args.resume else set()

    logging.info(f"  Questions:    {len(df)}")
    logging.info(f"  Judge labels: {len(judge)}")
    resume_info = ""
    if args.resume:
        resume_info = f" ({args.resume} mode)"
    logging.info(f"  Resumed:      {len(completed)} completed calls{resume_info}")
    logging.info(f"  Total calls:  {len(df) * len(VARIANTS)} planned\n")

    # ── Build task list ──────────────────────────────────────────
    logging.info("  Building task list...")
    tasks_list = []
    task_idx   = 0

    for _, row in df.iterrows():
        qid   = str(row["id"]).strip()
        gt    = row["ground_truth"]
        # Use row's subcategory if available (accept both "Subcat" and "subcategory"), else use CLI arg
        row_subcat_raw = row.get("subcategory") or row.get("Subcat") or subcat
        # Handle NaN/None — fall back to CLI arg
        if pd.isna(row_subcat_raw) or str(row_subcat_raw).strip().lower() in ('nan', 'none'):
            row_subcat = subcat
        else:
            row_subcat = str(row_subcat_raw).strip()

        prompt = build_standard_prompt(
            problem_text=row["problem_text"] if "problem_text" in row and pd.notna(row["problem_text"]) else "",
            problem_latex=row["problem_latex"],
            subcat=row_subcat
        )

        if (qid, "standard") in completed:
            continue
        tasks_list.append((task_idx, qid, prompt, gt, row["problem_latex"], row["answer_latex"], row_subcat))
        task_idx += 1

    # Apply limit if specified
    if args.limit and args.limit > 0:
        tasks_list = tasks_list[:args.limit]

    # ── Execute in parallel ──────────────────────────────────────
    max_workers = 8  # hardcoded default
    logging.info(f"  Executing {len(tasks_list)} calls ({max_workers} workers)...\n")

    parallel_exec = ParallelExecutor(
        client=client,
        results_file=results_file,
        extract_fn=extract_fn,
        max_workers=max_workers,
        rate_limit=1.0,  # hardcoded default
        max_tokens=max_tokens,
        ceiling=ceiling,
    )

    if len(tasks_list) > 0:
        pbar = tqdm(total=len(tasks_list), desc="API Calls", unit="call")
        parallel_exec.run(tasks_list, pbar_callback=lambda n: pbar.update(n))
        pbar.close()
    else:
        logging.info("  No new tasks (all completed, --resume in effect)")

    # ── Print quick results summary to stdout ─────────────────────
    print_results_summary(results_file)

    # ── Deduplicate results (if --resume created duplicates) ─────
    deduplicate_results(results_file)

    # ── Generate Stage 2 compatible failures JSONL ──────────────
    _generate_stage2_failures(results_file, output_dir, args.model, subcat)

    # ── LLM equivalence check for extraction failures (auto-enabled) ────────────
    if not args.no_equivalence_check and not args.dry_run:
        print("\n" + "─" * 60)
        print("  LLM equivalence check for extraction failures...")
        api_key = os.environ.get("OPENROUTER_API_KEY", "")
        rerun_equivalence_check(results_file, api_key=api_key)

    # ── Analysis ─────────────────────────────────────────────────
    logging.info("\n" + "─" * 60)
    logging.info("  Running analysis...")
    run_analysis(results_file, judge, args.model, output_dir)

    # Try to run summarize.py if it exists
    try:
        import summarize as sm
        sm.update_accuracy_file(args.input, args.model, results_file)
    except ImportError:
        pass  # summarize.py not required

    logging.info(f"\n  Raw results:     {results_file}")
    if len(tasks_list) > 0:
        logging.info(f"  Total API calls: {parallel_exec.call_count}")
    else:
        logging.info(f"  (Analyzed existing results, no new API calls)")



def _generate_stage2_failures(results_file: str, output_dir: str, model_name: str, subcat: str):
    """Generate Stage 2 compatible failures JSONL from unified inference output.
    
    Converts unified JSONL schema → Stage 2 (pipeline/label_judge.py) schema.
    Only outputs failures (correct != True).
    """
    MODEL_MAP = {cfg['model_id']: name for name, cfg in MODELS.items()}
    # Add alias for legacy 'o1' without prefix
    if 'o1' not in MODEL_MAP and 'openai/o1' in MODEL_MAP:
        MODEL_MAP['o1'] = MODEL_MAP['openai/o1']

    records = []
    with open(results_file) as f:
        for line in f:
            if line.strip():
                try:
                    records.append(json.loads(line))
                except json.JSONDecodeError:
                    pass

    if not records:
        return

    # Deduplicate: keep last per (question_id, variant)
    deduped, order = {}, {}
    for i, r in enumerate(records):
        key = (r['question_id'], r.get("variant", "standard"))
        deduped[key] = r
        order[key] = i

    out = []
    for key in sorted(deduped.keys(), key=lambda k: order[k]):
        r = deduped[key]
        if r.get('variant') != 'standard':
            continue
        if r.get('correct') is True:
            continue

        response = r.get('response', '') or ''
        extracted = r.get('extracted_answer')

        # Classify completion status
        if extracted is not None and extracted != '':
            comp_status = 'COMPLETE'
        elif not response.strip():
            comp_status = 'API_ERROR'
        elif len(response) >= 32760:
            comp_status = 'HARD_TRUNCATION'
        else:
            last = response[-400:]
            clean = bool(re.search(r'[.!?)\]}]\s*$', last)) or bool(re.search(r'(?:thus|therefore|in conclusion|hope this)', last, re.IGNORECASE))
            comp_status = 'ABANDONMENT' if clean else 'SOFT_TRUNCATION'

        out.append({
            'Problem_ID': r['question_id'],
            'Model': MODEL_MAP.get(r['model'], r['model']),
            'Subcat': subcat,
            'problem_latex': r.get('problem_latex', ''),
            'answer_latex': str(r.get('answer_latex', '')),
            'response': response,
            'ground_truth': r.get('ground_truth'),
            'extracted_answer': extracted,
            'response_correct': r.get('correct', False),
            'completion_status': comp_status,
            'Correct': 'FAIL',
            'format': r.get('format', 'latex'),
        })

    if out:
        outpath = os.path.join(output_dir, f'{model_name}_{subcat}_failures.jsonl')
        with open(outpath, 'w') as f:
            for rec in out:
                f.write(json.dumps(rec, default=str) + '\n')
        logging.info(f"  Generated Stage 2 output: {outpath} ({len(out)} failures)")

    # ── Generate professional-grade Excel with ALL records ─────────────
    _export_results_excel(deduped, order, output_dir, model_name, subcat, MODEL_MAP)


def _export_results_excel(deduped: dict, order: dict, output_dir: str,
                           model_name: str, subcat: str, model_map: dict):
    """Export ALL records to professional-grade Excel with formatted columns.

    Produces a comprehensive results workbook with:
      • All records (correct + incorrect, all variants)
      • Computed columns (response_correct, completion_status, Correct/FAIL)
      • Summary sheet with per-variant accuracy breakdown
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logging.warning("  openpyxl not available — skipping Excel export")
        return

    # Sort records by original order
    all_records = [deduped[k] for k in sorted(deduped, key=lambda k: order[k])]

    if not all_records:
        return

    MODEL_MAP = model_map

    # Build rows with all columns
    COLUMNS = [
        'Problem_ID', 'Model', 'Subcat', 'Variant', 'Correct',
        'response_correct', 'completion_status', 'extracted_answer',
        'ground_truth', 'answer_latex', 'problem_latex', 'response',
        'format', 'boxed_count', 'tokens_used', 'latency_ms',
        'finish_reason', 'error', 'timestamp',
    ]

    def classify_status(r):
        extracted = r.get('extracted_answer')
        response = r.get('response', '') or ''
        if extracted is not None and extracted != '':
            return 'COMPLETE'
        if not response.strip():
            return 'API_ERROR'
        if len(response) >= 32760:
            return 'HARD_TRUNCATION'
        last = response[-400:]
        clean = bool(re.search(r'[.!?)\]}]\s*$', last)) or bool(
            re.search(r'(?:thus|therefore|in conclusion|hope this)', last, re.IGNORECASE))
        return 'ABANDONMENT' if clean else 'SOFT_TRUNCATION'

    rows = []
    for r in all_records:
        correct_val = r.get('correct')
        if correct_val is None:
            correct_label = 'EXTRACT_FAIL'
        elif correct_val:
            correct_label = 'CORRECT'
        else:
            correct_label = 'FAIL'

        rows.append({
            'Problem_ID':       r.get('question_id', ''),
            'Model':            MODEL_MAP.get(r.get('model', ''), r.get('model', '')),
            'Subcat':           subcat,
            'Variant':          r.get('variant', ''),
            'Correct':          correct_label,
            'response_correct': correct_val if correct_val is not None else False,
            'completion_status': classify_status(r),
            'extracted_answer': str(r.get('extracted_answer', '')) if r.get('extracted_answer') is not None else '',
            'ground_truth':     str(r.get('ground_truth', '')),
            'answer_latex':     r.get('answer_latex', ''),
            'problem_latex':    r.get('problem_latex', ''),
            'response':         r.get('response', ''),
            'format':           r.get('format', 'latex'),
            'boxed_count':      r.get('boxed_count', 0),
            'tokens_used':      r.get('tokens_used', 0),
            'latency_ms':       r.get('latency_ms', 0),
            'finish_reason':    r.get('finish_reason', ''),
            'error':            r.get('error', ''),
            'timestamp':        r.get('timestamp', ''),
        })

    # ── Create workbook ──────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # Styles
    TITLE_FONT = Font(name='Calibri', bold=True, size=14, color='1F3864')
    HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    HEADER_FILL = PatternFill('solid', fgColor='1F3864')
    DATA_FONT = Font(name='Calibri', size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    CORRECT_FILL = PatternFill('solid', fgColor='C6EFCE')
    FAIL_FILL = PatternFill('solid', fgColor='FFC7CE')
    OTHER_FILL = PatternFill('solid', fgColor='FFEB9C')
    ALT_ROW = PatternFill('solid', fgColor='F2F2F2')

    # ── Sheet 1: All Results ─────────────────────────────────────────
    ws = wb.active
    ws.title = 'All_Results'
    ws.sheet_properties.tabColor = '1F3864'

    # Title
    ws.merge_cells('A1:T1')
    ws['A1'].value = f'Results: {model_name} — {subcat} '
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    # Headers
    for c, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=2, column=c, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Data rows
    for i, row_data in enumerate(rows):
        r_num = i + 3
        for c, col in enumerate(COLUMNS, 1):
            val = row_data.get(col, '')
            cell = ws.cell(row=r_num, column=c, value=val)
            cell.font = DATA_FONT
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if i % 2 == 1:
                cell.fill = ALT_ROW

        # Color the "Correct" column
        status = row_data.get('Correct', '')
        cell_c = ws.cell(row=r_num, column=5)
        if status == 'CORRECT':
            cell_c.fill = CORRECT_FILL
        elif status == 'FAIL':
            cell_c.fill = FAIL_FILL
        else:
            cell_c.fill = OTHER_FILL

    # Column widths
    col_widths = {
        'A': 22, 'B': 18, 'C': 12, 'D': 22, 'E': 14,
        'F': 16, 'G': 18, 'H': 16, 'I': 14, 'J': 30,
        'K': 40, 'L': 60, 'M': 10, 'N': 12, 'O': 12,
        'P': 12, 'Q': 14, 'R': 14, 'S': 22,
    }
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width

    # Row heights (data rows)
    for r_num in range(3, 3 + len(rows)):
        ws.row_dimensions[r_num].height = 60

    ws.freeze_panes = 'C3'
    ws.auto_filter.ref = f'A2:T{2 + len(rows)}'

    # ── Sheet 2: Summary ──────────────────────────────────────────────
    ws2 = wb.create_sheet('Summary')
    ws2.sheet_properties.tabColor = '59A14F'
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 14

    ws2.merge_cells('A1:C1')
    ws2['A1'].value = f'Summary — {model_name} ({subcat})'
    ws2['A1'].font = TITLE_FONT
    ws2.row_dimensions[1].height = 28

    # Stats
    total = len(rows)
    correct_count = sum(1 for r in rows if r['Correct'] == 'CORRECT')
    fail_count = sum(1 for r in rows if r['Correct'] == 'FAIL')
    other_count = sum(1 for r in rows if r['Correct'] not in ('CORRECT', 'FAIL'))

    # Per-variant breakdown
    variant_stats = {}
    for r in rows:
        v = r['Variant']
        if v not in variant_stats:
            variant_stats[v] = {'total': 0, 'correct': 0}
        variant_stats[v]['total'] += 1
        if r['Correct'] == 'CORRECT':
            variant_stats[v]['correct'] += 1

    r = 3
    ws2.merge_cells(f'A{r}:C{r}')
    ws2.cell(row=r, column=1, value='Overall').font = Font(name='Calibri', bold=True, size=12, color='1F3864')
    r = 4
    for h, w in [('Metric', 30), ('Count', 14), ('Percentage', 14)]:
        ws2.cell(row=r, column=['Metric', 'Count', 'Percentage'].index(h) + 1, value=h)
        ws2.cell(row=r, column=['Metric', 'Count', 'Percentage'].index(h) + 1).font = HEADER_FONT
        ws2.cell(row=r, column=['Metric', 'Count', 'Percentage'].index(h) + 1).fill = HEADER_FILL
        ws2.cell(row=r, column=['Metric', 'Count', 'Percentage'].index(h) + 1).border = thin_border
        ws2.cell(row=r, column=['Metric', 'Count', 'Percentage'].index(h) + 1).alignment = Alignment(horizontal='center')

    metrics = [
        ('Total Records', total, 1.0),
        ('Correct', correct_count, correct_count / total if total else 0),
        ('Failed', fail_count, fail_count / total if total else 0),
        ('Other (extract fail)', other_count, other_count / total if total else 0),
    ]
    for i, (metric, cnt, pct) in enumerate(metrics):
        r += 1
        ws2.cell(row=r, column=1, value=metric).font = DATA_FONT
        ws2.cell(row=r, column=2, value=cnt).font = DATA_FONT
        ws2.cell(row=r, column=2).number_format = '#,##0'
        ws2.cell(row=r, column=3, value=pct).font = DATA_FONT
        ws2.cell(row=r, column=3).number_format = '0.0%'
        for c in range(1, 4):
            ws2.cell(row=r, column=c).border = thin_border
            ws2.cell(row=r, column=c).alignment = Alignment(horizontal='center')
            if i % 2 == 0:
                ws2.cell(row=r, column=c).fill = ALT_ROW

    # Variant breakdown
    r += 2
    ws2.merge_cells(f'A{r}:C{r}')
    ws2.cell(row=r, column=1, value='By Variant').font = Font(name='Calibri', bold=True, size=12, color='1F3864')
    r += 1
    for h in ['Variant', 'Correct/Total', 'Accuracy']:
        ws2.cell(row=r, column=['Variant', 'Correct/Total', 'Accuracy'].index(h) + 1, value=h)
        ws2.cell(row=r, column=['Variant', 'Correct/Total', 'Accuracy'].index(h) + 1).font = HEADER_FONT
        ws2.cell(row=r, column=['Variant', 'Correct/Total', 'Accuracy'].index(h) + 1).fill = HEADER_FILL
        ws2.cell(row=r, column=['Variant', 'Correct/Total', 'Accuracy'].index(h) + 1).border = thin_border
        ws2.cell(row=r, column=['Variant', 'Correct/Total', 'Accuracy'].index(h) + 1).alignment = Alignment(horizontal='center')

    for i, (v, stats) in enumerate(variant_stats.items()):
        r += 1
        acc = stats['correct'] / stats['total'] if stats['total'] else 0
        ws2.cell(row=r, column=1, value=v).font = DATA_FONT
        ws2.cell(row=r, column=2, value=f"{stats['correct']}/{stats['total']}").font = DATA_FONT
        ws2.cell(row=r, column=3, value=acc).font = DATA_FONT
        ws2.cell(row=r, column=3).number_format = '0.0%'
        for c in range(1, 4):
            ws2.cell(row=r, column=c).border = thin_border
            ws2.cell(row=r, column=c).alignment = Alignment(horizontal='center')
            if i % 2 == 0:
                ws2.cell(row=r, column=c).fill = ALT_ROW

    # Save
    # Save Excel with same base name as JSONL
    xlsx_path = os.path.join(output_dir, f'{model_name}_results.xlsx')
    wb.save(xlsx_path)
    logging.info(f"  Generated Excel output: {xlsx_path} ({total} records)")



# ═════════════════════════════════════════════════════════════════════════
# 11. CLI
# ═════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="LinAlg-Bench · Unified Inference — bare 0-shot across all subcategories",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Subcategories (inferred from input data):
  determinant   eigenvalue   rank   nullity   trace
  multiplication   matrix_power   matrix_vector   transpose

Backends (auto-detected from model config):
  OpenRouter   — most models via openrouter.ai
  OpenAI       — gpt-4o, gpt-5.2 via api.openai.com
  GenAI        — gemini-* models via google-genai SDK

Examples:
  python pipeline/inference.py --input data/sample_3x3.csv --model Llama-3.3-70B --dry-run
  python pipeline/inference.py --input data/rank/rank_problems.csv --model GPT-4o
  python pipeline/inference.py --input data/eigen/eigen_problems.csv --model DeepSeek-V3 --resume all
  python pipeline/inference.py --input data/det/det_problems.csv --model GPT-5.2 --max-tokens 32768
        """,
    )
    parser.add_argument(
        "--model", default="Llama-3.3-70B",
        choices=list(MODELS.keys()),
        help="Model name from MODELS registry (default: Llama-3.3-70B)",
    )
    parser.add_argument(
        "--input", default=None,
        help="Path to input dataset (.xlsx or .csv). Default depends on subcat.",
    )
    parser.add_argument(
        "--judge", default=None,
        help="Forensic judge labels file (.xlsx or .csv)",
    )
    parser.add_argument(
        "--output", default=None,
        help="Output directory (default: results/{subcat}/{model})",
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview prompts without making API calls")
    parser.add_argument(
        "--resume", nargs='?', const='all', default=None,
        choices=['all', 'failures'],
        help="Resume: 'all' skip all records, 'failures' redo extraction failures")
    parser.add_argument("--seed", type=int, default=None,
                        help="Random seed for reproducibility (default: None)")
    parser.add_argument("--max-tokens", type=int, default=None,
                        help="Starting token budget (default: model min_tokens from config)")
    parser.add_argument("--api-base", default=None,
                        help="Override API base URL")
    parser.add_argument("--api-key-env", default=None,
                        help="Override environment variable name for API key")
    parser.add_argument("--no-equivalence-check", action="store_true",
                        help="Skip LLM equivalence check for extraction failures (default: auto-check enabled)")
    parser.add_argument("-n", "--limit", type=int, default=None,
                        help="Limit number of problems to process (default: all)")

    args = parser.parse_args()

    try:
        run_pipeline(args)
    except Exception as e:
        print(f"\n❌ FATAL ERROR: {type(e).__name__}: {e}", flush=True)
        import traceback
        traceback.print_exc()
        exit(1)


if __name__ == "__main__":
    main()


