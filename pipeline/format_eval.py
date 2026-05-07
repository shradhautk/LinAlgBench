"""
matrix_eval_pipeline.py
=======================
End-to-end resumable pipeline that:

  1. Ingests heterogeneous CSV / XLSX dumps from a folder, normalises columns
     via an alias table, and coalesces matrix / scalar ground-truth fields
     into the canonical evaluation columns.
  2. Sends each row through an LLM verification agent — but only when the
     existing `is_correct` label is False or missing (boolean gate).
  3. Persists evaluations to a JSONL checkpoint so re-runs can resume.
  4. Splits results by instruction format (ascii / latex / list) and writes
     a 4-sheet Excel report with per-format raw data, an accuracy pivot
     table grouped by cognitive category, and embedded matplotlib charts.

Usage
-----
    python matrix_eval_pipeline.py
    python matrix_eval_pipeline.py --data-dir data/full --concurrency 8 --resume
    python matrix_eval_pipeline.py --model-num 2 --resume
    python matrix_eval_pipeline.py --resume --force        # re-eval, ignore cache
"""

from __future__ import annotations

import argparse
import asyncio
import hashlib
import io
import json
import logging
import os
import re
import sys
import tempfile
import threading
from datetime import datetime
from textwrap import dedent
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

load_dotenv()

try:
    from tqdm.asyncio import tqdm as atqdm
except ImportError:
    atqdm = None


# ============================================================================
# CONFIGURATION
# ============================================================================
DEFAULT_DATA_DIR    = "data/sample"
DEFAULT_TEMP_DIR    = "temp"
DEFAULT_RESULTS_DIR = "results"
DEFAULT_OUTPUT_XLSX = "matrix_eval_results.xlsx"
DEFAULT_CHECKPOINT  = "eval_checkpoint.jsonl"

# Canonical column schema. EVERY incoming CSV / XLSX is normalised to this set.
# Columns absent from a file are filled with empty strings.
CSV_COLUMNS: List[str] = [
    "question_id",
    "format",
    "sub_category",
    "model_name",
    "prompt",
    "model_response",
    "model_response_short",
    "answer_latex",          # primary evaluation reference
    "answer_value",
    "ground_matrix",         # fallback for answer_latex (matrix-shaped truths)
    "ground_scalar",         # fallback for answer_value (scalar truths)
    "extracted_answer",
    "extracted_scalar",
    "extracted_matrix",
    "model_correct_orig",    # the original label (True/False/None)
    "finish_reason",
    "prompt_tokens",
    "total_tokens",
    "latency_ms",
    "timestamp",
    "error",
]

# Pipeline output columns (added by `evaluate_dataframe`).
MODEL_CORRECT_COL     = "model_correct"
MODEL_CORRECT_INT_COL = "model_correct_int"
MODEL_CORRECT_EXPL    = "model_correct_explanation"

# Alias table: keys MUST be lower-case and DISTINCT (Python keeps only the
# last value when duplicates appear in a dict literal — that was a real bug
# in the previous version).  Each alias points to a CANONICAL column name.
COLUMN_ALIASES: Dict[str, str] = {
    # format
    "format_type":          "format",
    "fmt":                  "format",
    # sub-category
    "subcategory":          "sub_category",
    "subcat":               "sub_category",
    "category":             "sub_category",
    # prompt
    "instruction":          "prompt",
    "question":             "prompt",
    # model response — long form
    "response":             "model_response",
    "completion":           "model_response",
    "model_output":         "model_response",
    # model response — short / extracted
    "response_tail":        "model_response_short",
    "response_short":       "model_response_short",
    # original correctness label
    "is_correct":           "model_correct_orig",
    "correct":              "model_correct_orig",
    # token counts
    "tokens_used":          "total_tokens",
}

# Columns that should be COALESCED into a canonical column AFTER mapping:
# {fallback_column: target_column}.  When `target_column` is empty/NaN for
# a given row, the value of `fallback_column` is copied in.  This is much
# clearer than overloading the alias / collision-handling logic.
COALESCE_FALLBACKS: Dict[str, str] = {
    "ground_matrix": "answer_latex",
    "ground_scalar": "answer_value",
}

# Cognitive category mapping for the accuracy chart x-axis.
COGNITIVE_CATEGORY_MAP: Dict[str, str] = {
    "trace":   "Reading",
    "trans":   "Reading",
    "mult":    "Arithmetic",
    "matvec":  "Arithmetic",
    "rank":    "Sequential",
    "nullity": "Sequential",
    "pow2":    "Recursive",
    "eig":     "Recursive",
    "inv":     "Compositional",
    "det":     "Compositional",
}
COGNITIVE_CATEGORY_ORDER = [
    "Reading", "Arithmetic", "Sequential", "Recursive", "Compositional",
]


# ============================================================================
# LOGGING
# ============================================================================
def setup_logging(level: int = logging.INFO) -> logging.Logger:
    logging.basicConfig(
        level=level,
        format="%(asctime)s - %(levelname)s - %(message)s",
        force=True,
    )
    logging.getLogger("LiteLLM").setLevel(logging.WARNING)
    logging.getLogger("google").setLevel(logging.WARNING)
    return logging.getLogger(__name__)

logger = setup_logging()


# ============================================================================
# UTILITIES
# ============================================================================
def bool_to_int(value: Any) -> Optional[int]:
    """Map Python booleans to ints; everything else becomes None."""
    if value is True:
        return 1
    if value is False:
        return 0
    return None


def coerce_bool(value: Any) -> Optional[bool]:
    """
    Tolerantly parse various truthy/falsy representations.
    Recognises: bool, 0/1, 'true'/'false'/'yes'/'no'/'t'/'f' (case-insensitive).
    Returns None if the value is missing or unrecognised.
    """
    if value is True or value is False:
        return value
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if pd.isna(value):
            return None
        return bool(int(value))
    s = str(value).strip().lower()
    if s in ("true", "t", "yes", "y", "1"):
        return True
    if s in ("false", "f", "no", "n", "0"):
        return False
    return None


def is_blank(value: Any) -> bool:
    """True iff the value is None, NaN, or an empty/whitespace-only string."""
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    return str(value).strip() == "" or str(value).strip().lower() == "nan"


def truncate(text: str, max_len: int, label: str) -> str:
    if not text or len(text) <= max_len:
        return text or ""
    return f"[TRUNCATED_BEGINNING_{label}] " + text[-max_len:]


def ensure_dir(path: str) -> None:
    parent = os.path.dirname(os.path.abspath(path))
    if parent:
        os.makedirs(parent, exist_ok=True)


def row_key(qid: str, model: str, fmt: str) -> Tuple[str, str, str]:
    """Stable identifier for one (question, model, format) triple."""
    return (str(qid).strip(), str(model).strip(), str(fmt).strip().lower())


def make_session_id(qid: str, idx: int, source_file: str) -> str:
    src_hash = hashlib.sha1((source_file or "").encode("utf-8")).hexdigest()[:8]
    return f"eval_{src_hash}_{qid}_{idx}"


def atomic_write_text(path: str, text: str) -> None:
    """Write *text* to *path* atomically (via a tmp file in the same dir)."""
    ensure_dir(path)
    parent = os.path.dirname(os.path.abspath(path)) or "."
    fd, tmp = tempfile.mkstemp(prefix=".tmp_", dir=parent)
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as fh:
            fh.write(text)
        os.replace(tmp, path)
    except Exception:
        if os.path.exists(tmp):
            try:
                os.unlink(tmp)
            except OSError:
                pass
        raise


def save_dataframe_backup(df: pd.DataFrame, base_path_no_ext: str, *, label: str) -> None:
    """Persist *df* as `<base>.csv` + `<base>.jsonl` + `<base>.meta.json`."""
    ensure_dir(base_path_no_ext + ".csv")
    df.to_csv(base_path_no_ext + ".csv", index=False)

    jsonl_lines = []
    for rec in df.to_dict(orient="records"):
        cleaned = {
            k: (None if (isinstance(v, float) and pd.isna(v)) else v)
            for k, v in rec.items()
        }
        jsonl_lines.append(json.dumps(cleaned, ensure_ascii=False, default=str))
    atomic_write_text(base_path_no_ext + ".jsonl", "\n".join(jsonl_lines) + "\n")

    meta = {
        "label":      label,
        "rows":       int(len(df)),
        "columns":    list(df.columns),
        "written_at": datetime.now().isoformat(),
    }
    atomic_write_text(base_path_no_ext + ".meta.json", json.dumps(meta, indent=2))
    logger.info(f"  Backup [{label}] -> {base_path_no_ext}.csv / .jsonl / .meta.json ({len(df)} rows)")


# ============================================================================
# RESUMABLE CHECKPOINT
# ============================================================================
class EvaluationCheckpoint:
    """JSONL append-only log of (qid, model, format) -> verdict records."""

    def __init__(self, path: str):
        self.path = path
        self._lock = threading.Lock()
        ensure_dir(self.path)
        if not os.path.exists(self.path):
            open(self.path, "a", encoding="utf-8").close()

    def load(self) -> Dict[Tuple[str, str, str], dict]:
        cache: Dict[Tuple[str, str, str], dict] = {}
        if not os.path.exists(self.path):
            return cache
        bad = 0
        with open(self.path, "r", encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                except json.JSONDecodeError:
                    bad += 1
                    continue
                cache[row_key(
                    obj.get("question_id", ""),
                    obj.get("model_name", ""),
                    obj.get("format", ""),
                )] = obj
        if bad:
            logger.warning(f"  Checkpoint: skipped {bad} corrupt JSONL line(s)")
        return cache

    def append(self, *, question_id, model_name, fmt, is_correct, explanation) -> None:
        record = {
            "question_id": str(question_id),
            "model_name":  str(model_name),
            "format":      str(fmt),
            "is_correct":  is_correct,
            "explanation": str(explanation or ""),
            "ts":          datetime.now().isoformat(),
        }
        line = json.dumps(record, ensure_ascii=False, default=str)
        with self._lock:
            with open(self.path, "a", encoding="utf-8") as fh:
                fh.write(line + "\n")


# ============================================================================
# LLM AGENT — single source of truth for model selection
# ============================================================================
# This is the ONLY definition of get_model_for_agent in the file. The previous
# version had two: a local one (with modelnum 1..4) plus an import from
# `models`, and the import silently shadowed the local one. The local one was
# never used and the typo OPENAI_API_KEYx made one branch a permanent dead-end.

def get_model_for_agent(modelnum: int = 2):
    """
    Return a configured LiteLlm instance for the verification agent.

    Args:
        modelnum:
            1 -> openai/gpt-5-mini            (uses OPENAI_API_KEY)
            2 -> openrouter/deepseek-v3.2     (uses OPENROUTER_API_KEY)  [default]
            3 -> openrouter/openai/gpt-5-mini (uses OPENROUTER_API_KEY)
            4 -> openrouter/openai/gpt-oss-120b (uses OPENROUTER_API_KEY)
    """
    from google.adk.models.lite_llm import LiteLlm

    table = {
        1: ("openai/gpt-5-mini",                   "OPENAI_API_KEY"),
        2: ("openrouter/deepseek/deepseek-v3.2",   "OPENROUTER_API_KEY"),
        3: ("openrouter/openai/gpt-5-mini",        "OPENROUTER_API_KEY"),
        4: ("openrouter/openai/gpt-oss-120b",      "OPENROUTER_API_KEY"),
    }
    if modelnum not in table:
        raise ValueError(f"Invalid modelnum: {modelnum}. Use 1, 2, 3, or 4.")

    model_name, key_env = table[modelnum]
    api_key = os.environ.get(key_env)
    if not api_key:
        logger.warning(
            f"Environment variable {key_env!r} is not set — "
            f"model {model_name!r} will fail at call time."
        )
    return LiteLlm(model=model_name, api_key=api_key)


def _build_agent(model):
    """Construct the verification LlmAgent with the equivalence prompt."""
    from google.adk.agents import LlmAgent
    return LlmAgent(
        model=model,
        name="verification_agent",
        description=(
            "Verifies if a model response conceptually matches a correct "
            "mathematical answer."
        ),
        instruction=dedent("""
            YOU ARE A MATHEMATICAL EQUIVALENCE VERIFICATION AGENT

            Your task: Determine if a model response mathematically equals the correct answer.

            KEY PRINCIPLE: Two expressions are CORRECT if they represent the SAME mathematical
            value, regardless of representation format.

            UNIVERSAL EQUIVALENCE RULES:
            1. Numerical equivalence: |val1 - val2| < 0.0001
            2. Structural equivalence: Matching dimensions + all elements equivalent
            3. Form independence: Accept fractions, decimals, expanded, factored, etc.
            4. Precision handling: Rounding and floating-point variations OK if within tolerance
            5. Notation irrelevance: LaTeX, ASCII, plain text — all equivalent if same math
            6. Operations equivalence: Simplified vs unsimplified, factored vs expanded OK

            VERIFICATION ALGORITHM:
            Step 1 - EXTRACT: Identify data type and extract components
            Step 2 - COMPARE: Compare based on type using equivalence rules above
            Step 3 - VALIDATE: Double-check comparison accuracy
            Step 4 - DECIDE: Return true if equivalent, false if not

            EXAMPLES THAT SHOULD BE CORRECT:
            - Model: -1/67           vs Expected: -0.0149            -> CORRECT (within tolerance)
            - Model: [[1/2,0],[0,1]] vs Expected: [[0.5,0],[0,1.0]]  -> CORRECT
            - Model: x^2 + 3x + 2    vs Expected: (x+1)(x+2)         -> CORRECT
            - Model: 0.333           vs Expected: 1/3                -> CORRECT (within tolerance)

            RESPOND WITH JSON ONLY:
            {"is_correct": true,  "explanation": "brief reason"}
            or
            {"is_correct": false, "explanation": "brief reason"}
            No markdown fences, no extra text.
        """).strip(),
    )


# Lazy singleton — built on first call to `evaluate_dataframe`.
_agent_runner = None
_agent_runner_lock = threading.Lock()


def initialize_agent_runner(modelnum: int = 2) -> None:
    """Create the agent Runner once; subsequent calls are no-ops."""
    global _agent_runner
    if _agent_runner is not None:
        return
    with _agent_runner_lock:
        if _agent_runner is not None:
            return
        from google.adk.runners import Runner
        from google.adk.sessions import InMemorySessionService

        model = get_model_for_agent(modelnum)
        agent = _build_agent(model)
        _agent_runner = Runner(
            agent=agent,
            app_name="matrix_eval_pipeline",
            session_service=InMemorySessionService(),
            auto_create_session=True,
        )
        logger.info(f"Agent runner created (modelnum={modelnum}).")


async def call_verification_agent(
    model_response: str,
    correct_answer: str,
    session_id: str,
    *,
    max_response_chars: int = 20_000,
) -> dict:
    """Send one verification request to the agent; return the parsed verdict."""
    from google.genai import types

    model_response = truncate(model_response, max_response_chars, "RESP")
    correct_answer = truncate(correct_answer, max_response_chars, "ANS")

    payload = json.dumps({
        "model_response": model_response,
        "correct_answer": correct_answer,
    })
    user_content = types.Content(role="user", parts=[types.Part(text=payload)])

    response_text: Optional[str] = None
    async for event in _agent_runner.run_async(
        user_id="verifier",
        session_id=session_id,
        new_message=user_content,
    ):
        if event.is_final_response() and event.content and event.content.parts:
            response_text = event.content.parts[0].text
            break

    if not response_text:
        return {"is_correct": False, "explanation": "No response from agent"}

    text = response_text.strip()
    for fence in ("```json", "```"):
        if text.startswith(fence):
            text = text[len(fence):]
    if text.endswith("```"):
        text = text[:-3]
    text = text.strip()

    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return {"is_correct": False, "explanation": f"JSON parse failed: {text[:100]}"}


# ============================================================================
# STEP 1 — INGEST: column normalisation + multi-format file loading
# ============================================================================
_CANDIDATE_ENCODINGS = ("utf-8", "utf-8-sig", "cp1252", "latin-1")


def _read_csv_with_fallback(fpath: str, **kwargs) -> Tuple[pd.DataFrame, str]:
    """
    Try `pd.read_csv` with each candidate encoding in turn; return the
    DataFrame and the encoding that worked. `latin-1` is guaranteed to
    succeed because every byte sequence is valid latin-1.
    """
    last_exc: Optional[Exception] = None
    for enc in _CANDIDATE_ENCODINGS:
        try:
            return pd.read_csv(fpath, encoding=enc, **kwargs), enc
        except UnicodeDecodeError as exc:
            last_exc = exc
            continue
    # Should be unreachable since latin-1 cannot raise UnicodeDecodeError
    raise UnicodeDecodeError(  # pragma: no cover
        "all", b"", 0, 1, f"All encodings failed: {last_exc}"
    )


def _apply_column_mapping(
    df: pd.DataFrame,
    expected_cols: List[str],
    source_label: str,
) -> pd.DataFrame:
    """
    Rename incoming columns to the canonical schema.

    - Case-insensitive matching against canonical names then against aliases.
    - When two source columns map to the same canonical target, values are
      coalesced left-to-right (first non-empty wins).
    - Unknown columns are kept under the `_extra_` prefix for traceability.
    - Every canonical column is guaranteed to exist (filled with "" if absent).
    """
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    canonical_lower = {c.lower(): c for c in expected_cols}
    rename_plan: Dict[str, str] = {}

    for col in df.columns:
        cl = col.lower()
        if cl in canonical_lower:
            rename_plan[col] = canonical_lower[cl]
        elif cl in COLUMN_ALIASES:
            rename_plan[col] = COLUMN_ALIASES[cl]
        # else: leave unmapped — handled below as `_extra_*`

    # Group sources by the canonical target they map to.
    target_to_sources: Dict[str, List[str]] = {}
    for src, tgt in rename_plan.items():
        target_to_sources.setdefault(tgt, []).append(src)

    new_df = pd.DataFrame(index=df.index)
    consumed: set = set()

    for tgt, srcs in target_to_sources.items():
        # Preserve original column order for deterministic coalesce priority
        ordered = [s for s in df.columns if s in srcs]
        consumed.update(ordered)

        if len(ordered) == 1:
            new_df[tgt] = df[ordered[0]]
            continue

        # Multiple sources -> coalesce: keep first non-empty value per row
        logger.info(
            f"  [{source_label}] {ordered} -> '{tgt}' (coalesced left-to-right)"
        )
        series = df[ordered[0]]
        for extra in ordered[1:]:
            series = series.where(
                series.notna() & (series.astype(str).str.strip().str.len() > 0),
                df[extra],
            )
        new_df[tgt] = series

    # Carry forward any unmapped columns under `_extra_` prefix
    for col in df.columns:
        if col not in consumed:
            new_df[f"_extra_{col}"] = df[col]

    # Guarantee every canonical column exists
    for c in expected_cols:
        if c not in new_df.columns:
            new_df[c] = ""

    return new_df


def _coalesce_fallback_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Apply COALESCE_FALLBACKS: when `target` is blank, copy from `fallback`.
    Both columns are preserved; only the target is filled in.
    """
    df = df.copy()
    for fallback, target in COALESCE_FALLBACKS.items():
        if fallback not in df.columns or target not in df.columns:
            continue
        target_blank   = df[target].apply(is_blank)
        fallback_value = df[fallback].apply(lambda v: not is_blank(v))
        actually_fill  = target_blank & fallback_value

        n_actual = int(actually_fill.sum())
        if n_actual:
            df.loc[actually_fill, target] = df.loc[actually_fill, fallback]
            logger.info(
                f"  Coalesce: '{fallback}' -> '{target}' "
                f"({n_actual} value(s) copied)"
            )
    return df


def _sniff_and_load_csv(fpath: str, expected_cols: List[str]) -> pd.DataFrame:
    """Load a CSV, autodetecting whether row 0 is a header."""
    # Peek at the first row to decide header vs positional
    peek, enc = _read_csv_with_fallback(fpath, header=None, nrows=1, dtype=str)
    logger.debug(f"  '{os.path.basename(fpath)}': decoded as '{enc}'")
    first_row = {str(v).strip().lower() for v in peek.iloc[0]}

    canonical_lower = {c.lower() for c in expected_cols}
    alias_lower     = set(COLUMN_ALIASES.keys())
    overlap = len(first_row & (canonical_lower | alias_lower))
    has_header = overlap >= max(3, len(expected_cols) // 6)

    if has_header:
        df, _ = _read_csv_with_fallback(fpath, header=0, dtype=str)
        df = _apply_column_mapping(
            df, expected_cols, source_label=os.path.basename(fpath)
        )
        return df

    # Positional fallback (truly headerless dump)
    n_cols = peek.shape[1]
    if n_cols < len(expected_cols):
        raise ValueError(
            f"Headerless file '{fpath}' has only {n_cols} columns "
            f"(expected at least {len(expected_cols)})."
        )
    if n_cols > len(expected_cols):
        logger.warning(
            f"  '{os.path.basename(fpath)}': {n_cols} columns, "
            f"expected {len(expected_cols)} — extra columns dropped."
        )
    df, _ = _read_csv_with_fallback(fpath, header=None, dtype=str)
    df = df.iloc[:, :len(expected_cols)]
    df.columns = expected_cols
    return df


def _load_xlsx(fpath: str, expected_cols: List[str]) -> pd.DataFrame:
    """Load every sheet of an xlsx, mapping columns per-sheet."""
    sheets = pd.read_excel(fpath, sheet_name=None, dtype=str)
    frames = []
    for sheet_name, sdf in sheets.items():
        if sdf.empty:
            logger.info(f"    Sheet '{sheet_name}' is empty — skipped")
            continue
        sdf = _apply_column_mapping(
            sdf, expected_cols,
            source_label=f"{os.path.basename(fpath)}[{sheet_name}]",
        )
        sdf["_sheet"] = sheet_name
        frames.append(sdf)
        logger.info(f"    Sheet '{sheet_name}': {len(sdf)} rows")
    if not frames:
        raise ValueError(f"All sheets in '{fpath}' are empty.")
    return pd.concat(frames, ignore_index=True)


_CSV_EXTS  = {".csv"}
_XLSX_EXTS = {".xlsx", ".xlsm"}
_ALL_EXTS  = _CSV_EXTS | _XLSX_EXTS


def ingest_files(data_dir: str) -> pd.DataFrame:
    """
    Read every supported file in *data_dir* and stack into a master DataFrame.
    Applies alias mapping, fallback coalesce, and basic post-processing.
    """
    if not os.path.isdir(data_dir):
        raise FileNotFoundError(f"Input directory not found: {data_dir}")

    files = sorted(
        f for f in os.listdir(data_dir)
        if os.path.splitext(f.lower())[1] in _ALL_EXTS
    )
    if not files:
        raise FileNotFoundError(
            f"No CSV/XLSX files in '{data_dir}' "
            f"(supported: {sorted(_ALL_EXTS)})"
        )

    frames = []
    for fname in files:
        fpath = os.path.join(data_dir, fname)
        ext = os.path.splitext(fname.lower())[1]
        try:
            if ext in _XLSX_EXTS:
                df = _load_xlsx(fpath, CSV_COLUMNS)
                logger.info(f"  Loaded {len(df):>5} rows  <- {fname}  [Excel]")
            else:
                df = _sniff_and_load_csv(fpath, CSV_COLUMNS)
                logger.info(f"  Loaded {len(df):>5} rows  <- {fname}  [CSV]")
        except Exception as exc:
            logger.warning(f"  Skipping '{fname}': {exc}")
            continue

        df["_source_file"] = fname
        frames.append(df)

    if not frames:
        raise ValueError("All input files failed to load.")

    master = pd.concat(frames, ignore_index=True)

    # Apply ground_matrix -> answer_latex / ground_scalar -> answer_value coalesce
    master = _coalesce_fallback_columns(master)

    # Tolerantly normalise the original boolean label
    master["model_correct_orig"] = master["model_correct_orig"].apply(coerce_bool)

    # Trim string columns used as keys
    for col in ("question_id", "format", "sub_category", "model_name"):
        master[col] = master[col].fillna("").astype(str).str.strip()
    master["format"] = master["format"].str.lower()

    logger.info(f"Master DataFrame: {len(master)} rows x {len(master.columns)} columns")
    logger.info(f"  Formats : {master['format'].value_counts().to_dict()}")
    logger.info(f"  Sub-cats: {sorted(master['sub_category'].unique())}")
    logger.info(f"  Models  : {sorted(master['model_name'].unique())}")
    return master


# ============================================================================
# STEP 2 — EVALUATION (gated, concurrent, resumable)
# ============================================================================
async def _evaluate_one(
    *,
    pos: int,
    row: pd.Series,
    semaphore: asyncio.Semaphore,
    checkpoint: EvaluationCheckpoint,
    max_retries: int,
    base_backoff: float,
) -> Tuple[int, Optional[bool], str]:
    """Send one row through the agent with retries; persist to checkpoint."""
    qid    = str(row.get("question_id", "")).strip()
    fmt    = str(row.get("format", "")).strip()
    model  = str(row.get("model_name", "")).strip()
    source = str(row.get("_source_file", "")).strip()

    model_resp = "" if pd.isna(row.get("model_response")) else str(row["model_response"]).strip()
    truth      = "" if pd.isna(row.get("answer_latex"))   else str(row["answer_latex"]).strip()

    if not model_resp or not truth or model_resp.lower() == "nan":
        msg = "skipped — missing model_response or answer_latex"
        logger.warning(f"  Row {pos} ({qid}): {msg}")
        checkpoint.append(question_id=qid, model_name=model, fmt=fmt,
                          is_correct=None, explanation=msg)
        return pos, None, msg

    session_id = make_session_id(qid, pos, source)
    last_exc: Optional[Exception] = None

    async with semaphore:
        for attempt in range(1, max_retries + 1):
            try:
                result      = await call_verification_agent(model_resp, truth, session_id)
                is_correct  = bool(result.get("is_correct", False))
                explanation = str(result.get("explanation", ""))[:1000]
                checkpoint.append(question_id=qid, model_name=model, fmt=fmt,
                                  is_correct=is_correct, explanation=explanation)
                return pos, is_correct, explanation
            except Exception as exc:
                last_exc = exc
                if attempt < max_retries:
                    wait = base_backoff * (2 ** (attempt - 1))
                    logger.warning(
                        f"  Row {pos} ({qid}): attempt {attempt}/{max_retries} "
                        f"failed — {exc}; retrying in {wait:.1f}s"
                    )
                    await asyncio.sleep(wait)
                else:
                    logger.error(
                        f"  Row {pos} ({qid}): all {max_retries} attempts failed — {exc}"
                    )

    msg = f"agent error after {max_retries} attempts: {last_exc}"
    checkpoint.append(question_id=qid, model_name=model, fmt=fmt,
                      is_correct=False, explanation=msg)
    return pos, False, msg


async def evaluate_dataframe(
    df: pd.DataFrame,
    *,
    checkpoint_path: str,
    concurrency: int = 5,
    max_retries: int = 3,
    base_backoff: float = 1.0,
    resume: bool = False,
    force: bool = False,
    modelnum: int = 2,
) -> pd.DataFrame:
    """
    Evaluate every row, gated by the original `model_correct_orig` label:
      - True  -> trusted as-is (no LLM call)
      - False / None / unknown -> sent to the LLM verification agent

    Adds three preserved output columns:
      model_correct              True / False
      model_correct_int          1 / 0
      model_correct_explanation  agent reasoning (or gating message)
    """
    initialize_agent_runner(modelnum)
    checkpoint = EvaluationCheckpoint(checkpoint_path)

    cache: Dict[Tuple[str, str, str], dict] = {}
    if resume and not force:
        cache = checkpoint.load()
        logger.info(f"Resume: loaded {len(cache)} cached evaluation(s)")

    total = len(df)
    flags:        List[Optional[bool]] = [None] * total
    explanations: List[str]            = [""]   * total
    pending: List[Tuple[int, pd.Series]] = []

    n_trusted = n_cached = n_to_eval = 0

    # ── Routing pass: gate first, then cache, then queue for agent ──────────
    for pos, (_, row) in enumerate(df.iterrows()):
        orig = coerce_bool(row.get("model_correct_orig"))
        if orig is True and not force:
            flags[pos] = True
            explanations[pos] = "trusted from model_correct_orig=True (no agent call)"
            n_trusted += 1
            continue

        key = row_key(
            row.get("question_id", ""),
            row.get("model_name", ""),
            row.get("format", ""),
        )
        if not force and key in cache:
            cached = cache[key]
            flags[pos] = cached.get("is_correct")
            explanations[pos] = str(cached.get("explanation", ""))[:1000]
            n_cached += 1
            continue

        pending.append((pos, row))
        n_to_eval += 1

    logger.info(
        f"Routing: {n_trusted} trusted-True (gate), "
        f"{n_cached} cache hits, {n_to_eval} live agent calls "
        f"(total {total})"
    )

    # ── Agent pass: only the rows that actually need it ─────────────────────
    if pending:
        semaphore = asyncio.Semaphore(max(1, int(concurrency)))
        coros = [
            _evaluate_one(
                pos=pos, row=row, semaphore=semaphore, checkpoint=checkpoint,
                max_retries=max_retries, base_backoff=base_backoff,
            )
            for pos, row in pending
        ]
        iterator = (
            atqdm.as_completed(coros, total=len(coros), desc="evaluating")
            if atqdm is not None
            else asyncio.as_completed(coros)
        )
        done = 0
        for fut in iterator:
            pos, is_correct, expl = await fut
            flags[pos] = is_correct
            explanations[pos] = expl
            done += 1
            if atqdm is None and (done % 25 == 0 or done == len(pending)):
                logger.info(f"  progress: {done}/{len(pending)}")
    else:
        logger.info("No live agent calls needed (all rows handled by gate or cache).")

    out = df.copy()
    out[MODEL_CORRECT_COL]     = flags
    out[MODEL_CORRECT_INT_COL] = out[MODEL_CORRECT_COL].apply(bool_to_int)
    out[MODEL_CORRECT_EXPL]    = explanations

    evaluated = int(out[MODEL_CORRECT_INT_COL].notna().sum())
    correct   = int(out[MODEL_CORRECT_INT_COL].sum(skipna=True)) if evaluated else 0
    pct       = f"{100.0 * correct / evaluated:.1f}%" if evaluated else "N/A"
    logger.info(f"Evaluation complete — {correct}/{evaluated} correct ({pct})")
    return out


# ============================================================================
# STEP 3 — FORMAT SPLIT
# ============================================================================
def split_by_format(df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    """Partition rows by `format` column (already lower-cased in ingest)."""
    splits: Dict[str, pd.DataFrame] = {}
    fmts = df["format"].fillna("").astype(str).str.strip().str.lower()
    for fmt in ("ascii", "latex", "list"):
        subset = df[fmts == fmt].copy().reset_index(drop=True)
        splits[fmt] = subset
        logger.info(f"  Format '{fmt}': {len(subset)} rows")

    unknown_mask = ~fmts.isin({"ascii", "latex", "list"})
    if unknown_mask.any():
        unique = df.loc[unknown_mask, "format"].unique().tolist()
        logger.warning(
            f"  {int(unknown_mask.sum())} rows with unrecognised formats: {unique}"
        )
    return splits


# ============================================================================
# STEP 4 — ACCURACY PIVOT
# ============================================================================
def _extract_matrix_size(question_id: str) -> str:
    m = re.search(r"(\d+x\d+)", str(question_id), re.IGNORECASE)
    return m.group(1) if m else "unknown"


def _get_cognitive_category(problem_type: str) -> str:
    """'eig_4x4' -> 'Recursive', 'rank_5x5' -> 'Sequential', etc."""
    sub_cat = re.sub(r"_\d+x\d+$", "", problem_type, flags=re.IGNORECASE).lower()
    return COGNITIVE_CATEGORY_MAP.get(sub_cat, "Other")


def compute_accuracy(df: pd.DataFrame) -> dict:
    """Build the pivot structure consumed by `_write_pivot_sheet`.

    'totals' counts questions per problem_type for ONE (model, format) slice.
    Since every model is asked every question in every format, this equals
    the per-cell question count — which is exactly what the footer's
    "Questions analysed" row reports.
    """
    if df.empty or MODEL_CORRECT_INT_COL not in df.columns:
        return {}

    work = df.copy()
    work["matrix_size"]  = work["question_id"].apply(_extract_matrix_size)
    work["problem_type"] = work["sub_category"].str.strip() + "_" + work["matrix_size"]
    work = work[work[MODEL_CORRECT_INT_COL].notna()]
    if work.empty:
        return {}

    work["fmt"] = work["format"].str.strip().str.upper()

    models        = sorted(work["model_name"].unique().tolist())
    formats       = [f for f in ["ASCII", "LATEX", "LIST"] if f in work["fmt"].unique()]
    problem_types = sorted(work["problem_type"].unique().tolist())

    # Pivots: rows × (model, format) cells
    correct_pivot = work.pivot_table(
        index="problem_type", columns=["model_name", "fmt"],
        values=MODEL_CORRECT_INT_COL, aggfunc="sum", fill_value=0,
    ).reindex(problem_types, fill_value=0)

    total_pivot = work.pivot_table(
        index="problem_type", columns=["model_name", "fmt"],
        values=MODEL_CORRECT_INT_COL, aggfunc="count", fill_value=0,
    ).reindex(problem_types, fill_value=0)

    # 'totals' = question count per problem_type for ONE (model, format) cell.
    # Pick the first cell deterministically; falls back to global unique count
    # if the pivot has no columns (empty data edge case).
    if not total_pivot.columns.empty:
        first_cell = total_pivot.columns[0]
        totals = total_pivot[first_cell].copy()
    else:
        totals = (
            work.groupby("problem_type")["question_id"]
            .nunique()
            .reindex(problem_types, fill_value=0)
        )

    return dict(
        problem_types=problem_types,
        models=models,
        formats=formats,
        totals=totals,
        correct_pivot=correct_pivot,
        total_pivot=total_pivot,
    )

# ============================================================================
# STEP 5 — STYLING + EXCEL EXPORT
# ============================================================================
def _style_data_sheet(ws, df: pd.DataFrame) -> None:
    """Apply consistent styling to a raw-data sheet."""
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    data_font   = Font(name="Arial", size=10)
    center      = Alignment(horizontal="center", vertical="top", wrap_text=False)
    left        = Alignment(horizontal="left",   vertical="top", wrap_text=False)
    true_fill   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    false_fill  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    bool_cols = {MODEL_CORRECT_COL, MODEL_CORRECT_INT_COL, "model_correct_orig"}
    bool_idxs = {i + 1 for i, c in enumerate(df.columns) if c in bool_cols}
    n_cols = len(df.columns)

    for r_idx, row_cells in enumerate(
        ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=n_cols), start=1
    ):
        for c_idx, cell in enumerate(row_cells, start=1):
            if r_idx == 1:
                cell.font, cell.fill, cell.alignment = header_font, header_fill, center
            else:
                cell.font = data_font
                cell.alignment = center if c_idx in bool_idxs else left
                if c_idx in bool_idxs and cell.value is not None:
                    if cell.value in (True, 1, "True", "TRUE"):
                        cell.fill = true_fill
                    elif cell.value in (False, 0, "False", "FALSE"):
                        cell.fill = false_fill

    for c_idx, col_cells in enumerate(
        ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=n_cols), start=1
    ):
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=10,
        )
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 3, 60)

    ws.freeze_panes = "A2"


def _write_pivot_sheet(ws, pivot: dict) -> int:
    """Render the accuracy pivot table; return the next free row."""
    if not pivot:
        ws.cell(row=1, column=1, value="No data available")
        return 3

    problem_types = pivot["problem_types"]
    models        = pivot["models"]
    formats       = pivot["formats"]
    totals        = pivot["totals"]
    correct_pivot = pivot["correct_pivot"]
    total_pivot   = pivot["total_pivot"]

    nf, nm = len(formats), len(models)

    def _fill(color):
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def _font(bold=False, color="000000", size=10):
        return Font(name="Arial", bold=bold, color=color, size=size)

    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cnt    = Alignment(horizontal="center", vertical="center")
    lft    = Alignment(horizontal="left",   vertical="center")

    DARK_BLUE, MID_BLUE = "1F4E79", "2E75B6"
    ROW_BLUE, ROW_WHITE, ROW_CREAM = "D9E1F2", "FFFFFF", "FFF2CC"
    FOOTER_BLUE = "4472C4"
    GREEN, AMBER, RED = "C6EFCE", "FFEB9C", "FFC7CE"

    def _cell(row, col, value, *, fill=None, bold=False, fcolor="000000", align=None):
        c = ws.cell(row=row, column=col, value=value)
        if fill:
            c.fill = _fill(fill)
        c.font = _font(bold=bold, color=fcolor)
        c.alignment = align or cnt
        c.border = border
        return c

    # Header rows
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    _cell(1, 1, "Problem Type", fill=DARK_BLUE, bold=True, fcolor="FFFFFF")
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    _cell(1, 2, "Total", fill=DARK_BLUE, bold=True, fcolor="FFFFFF")
    for mi, model in enumerate(models):
        cs = 3 + mi * nf
        ce = cs + nf - 1
        if cs < ce:
            ws.merge_cells(start_row=1, start_column=cs, end_row=1, end_column=ce)
        _cell(1, cs, model, fill=MID_BLUE, bold=True, fcolor="FFFFFF")
    for mi in range(nm):
        for fi, fmt in enumerate(formats):
            _cell(2, 3 + mi * nf + fi, fmt,
                  fill=MID_BLUE, bold=True, fcolor="FFFFFF")

    # Data rows
    for ri, pt in enumerate(problem_types):
        row = 3 + ri
        row_fill = ROW_CREAM if ri % 2 else ROW_WHITE
        _cell(row, 1, pt, fill=ROW_BLUE, bold=True, align=lft)
        _cell(row, 2, int(totals.get(pt, 0)), fill=ROW_BLUE, bold=True)
        for mi, model in enumerate(models):
            for fi, fmt in enumerate(formats):
                col = 3 + mi * nf + fi
                try:
                    val = int(correct_pivot.loc[pt, (model, fmt)])
                except KeyError:
                    val = 0
                _cell(row, col, val, fill=row_fill)

    # Footer rows
    foot = 3 + len(problem_types)
    _cell(foot, 1, "Questions analysed",
          fill=FOOTER_BLUE, bold=True, fcolor="FFFFFF", align=lft)
    _cell(foot, 2, int(totals.sum()),
          fill=FOOTER_BLUE, bold=True, fcolor="FFFFFF")
    for mi, model in enumerate(models):
        for fi, fmt in enumerate(formats):
            col = 3 + mi * nf + fi
            try:
                val = int(total_pivot[(model, fmt)].sum())
            except KeyError:
                val = 0
            _cell(foot, col, val, fill=FOOTER_BLUE, bold=True, fcolor="FFFFFF")

    _cell(foot + 1, 1, "Correct (incl. apparent)",
          fill=FOOTER_BLUE, bold=True, fcolor="FFFFFF", align=lft)
    grand_correct = 0
    for mi, model in enumerate(models):
        for fi, fmt in enumerate(formats):
            col = 3 + mi * nf + fi
            try:
                val = int(correct_pivot[(model, fmt)].sum())
            except KeyError:
                val = 0
            grand_correct += val
            _cell(foot + 1, col, val, fill=FOOTER_BLUE, bold=True, fcolor="FFFFFF")
    _cell(foot + 1, 2, "",
          fill=FOOTER_BLUE, bold=True, fcolor="FFFFFF")

    _cell(foot + 2, 1, "Accuracy",
          fill=FOOTER_BLUE, bold=True, fcolor="FFFFFF", align=lft)
    grand_q = grand_c = 0
    for mi, model in enumerate(models):
        for fi, fmt in enumerate(formats):
            col = 3 + mi * nf + fi
            try:
                q = int(total_pivot[(model, fmt)].sum())
                c = int(correct_pivot[(model, fmt)].sum())
            except KeyError:
                q = c = 0
            grand_q += q
            grand_c += c
            pct = round(c / q * 100, 2) if q else 0.0
            band = GREEN if pct >= 75 else (AMBER if pct >= 50 else RED)
            _cell(foot + 2, col, f"{pct:.2f}%", fill=band, bold=True)

    overall = round(grand_c / grand_q * 100, 2) if grand_q else 0.0
    band = GREEN if overall >= 75 else (AMBER if overall >= 50 else RED)
    _cell(foot + 2, 2, "", fill=band, bold=True)

    # Layout
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    for mi in range(nm):
        for fi in range(nf):
            ws.column_dimensions[get_column_letter(3 + mi * nf + fi)].width = 12
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    for ri in range(len(problem_types) + 3):
        ws.row_dimensions[3 + ri].height = 16
    ws.freeze_panes = "A3"

    return foot + 5


# ============================================================================
# STEP 5b — CHARTS (matplotlib, embedded as PNG)
# ============================================================================
def _generate_matrix_chart(pivot: dict, matrix_size: str) -> Optional[io.BytesIO]:
    """Grouped bar chart for one matrix size. None if no data for that size."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    problem_types = pivot["problem_types"]
    models        = pivot["models"]
    formats       = pivot["formats"]
    correct_pivot = pivot["correct_pivot"]
    total_pivot   = pivot["total_pivot"]

    tag = f"_{matrix_size.lower()}"
    size_pts = sorted(pt for pt in problem_types if pt.lower().endswith(tag))
    if not size_pts:
        return None

    fmt_order  = [f for f in ["LATEX", "LIST", "ASCII"] if f in formats]
    fmt_colors = {"LATEX": "#4472C4", "LIST": "#ED7D31", "ASCII": "#70AD47"}
    fmt_labels = {"LATEX": "LaTeX", "LIST": "List",   "ASCII": "ASCII"}

    cat_data: Dict[str, Dict[str, List[int]]] = {
        cat: {fmt: [0, 0] for fmt in fmt_order}
        for cat in COGNITIVE_CATEGORY_ORDER
    }
    for pt in size_pts:
        cat = _get_cognitive_category(pt)
        if cat not in cat_data:
            continue
        for fmt in fmt_order:
            for model in models:
                try:
                    cat_data[cat][fmt][0] += int(correct_pivot.loc[pt, (model, fmt)])
                    cat_data[cat][fmt][1] += int(total_pivot.loc[pt, (model, fmt)])
                except KeyError:
                    pass

    present = [c for c in COGNITIVE_CATEGORY_ORDER
               if any(cat_data[c][f][1] > 0 for f in fmt_order)]
    if not present:
        return None

    fmt_acc = {
        fmt: [
            round(cat_data[c][fmt][0] / cat_data[c][fmt][1] * 100)
            if cat_data[c][fmt][1] else 0
            for c in present
        ]
        for fmt in fmt_order
    }

    n_groups, n_fmts = len(present), len(fmt_order)
    x = np.arange(n_groups)
    width = 0.22

    fig, ax = plt.subplots(figsize=(max(6, n_groups * 1.6), 5))
    fig.patch.set_facecolor("white")
    for i, fmt in enumerate(fmt_order):
        offset = (i - (n_fmts - 1) / 2) * width
        bars = ax.bar(x + offset, fmt_acc[fmt], width,
                      color=fmt_colors[fmt], label=fmt_labels[fmt], zorder=3)
        for bar, val in zip(bars, fmt_acc[fmt]):
            ax.text(bar.get_x() + bar.get_width() / 2,
                    bar.get_height() + 1.5, str(val),
                    ha="center", va="bottom", fontsize=8, fontweight="bold")

    ax.set_xticks(x)
    ax.set_xticklabels(present, fontsize=9)
    ax.set_ylabel("Average accuracy (%)", fontsize=9)
    ax.set_ylim(0, 118)
    ax.set_title(f"{matrix_size} Matrices: Format Accuracy by Cognitive Level",
                 fontsize=11, pad=12)
    ax.legend(loc="upper right", fontsize=8, frameon=False)
    ax.set_axisbelow(True)
    ax.yaxis.grid(True, color="#E0E0E0", linewidth=0.7)
    for spine in ("top", "right", "left"):
        ax.spines[spine].set_visible(False)
    ax.tick_params(left=False)

    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf


def _insert_charts(ws, pivot: dict, start_row: int) -> None:
    """Embed one chart per detected matrix size, side-by-side."""
    from openpyxl.drawing.image import Image as XLImage

    IMG_W, IMG_H, COL_SPAN = 480, 300, 9
    sizes = sorted(
        {
            re.search(r"(\d+x\d+)", pt, re.IGNORECASE).group(1)
            for pt in pivot["problem_types"]
            if re.search(r"(\d+x\d+)", pt, re.IGNORECASE)
        },
        key=lambda s: int(s.split("x")[0]),
    )
    for idx, size in enumerate(sizes):
        buf = _generate_matrix_chart(pivot, size)
        if buf is None:
            continue
        img = XLImage(buf)
        img.width, img.height = IMG_W, IMG_H
        col = get_column_letter(1 + idx * COL_SPAN)
        ws.add_image(img, f"{col}{start_row}")
    for r in range(start_row, start_row + 20):
        ws.row_dimensions[r].height = 15


def export_to_excel(
    format_splits: Dict[str, pd.DataFrame],
    accuracy_pivot: dict,
    output_path: str,
) -> None:
    """Write the four-sheet Excel report (3 raw + 1 accuracy)."""
    ensure_dir(output_path)

    data_sheets = {
        "ascii_data": format_splits.get("ascii", pd.DataFrame()),
        "latex_data": format_splits.get("latex", pd.DataFrame()),
        "list_data":  format_splits.get("list",  pd.DataFrame()),
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in data_sheets.items():
            (df if not df.empty
             else pd.DataFrame([{"info": f"No data for {sheet_name}"}])
             ).to_excel(writer, sheet_name=sheet_name, index=False)
        pd.DataFrame().to_excel(writer, sheet_name="accuracy_results", index=False)

    wb = load_workbook(output_path)
    for sheet_name, df in data_sheets.items():
        styling_df = df if not df.empty else pd.DataFrame([{"info": ""}])
        _style_data_sheet(wb[sheet_name], styling_df)

    ws_acc = wb["accuracy_results"]
    chart_row = _write_pivot_sheet(ws_acc, accuracy_pivot)
    if accuracy_pivot:
        _insert_charts(ws_acc, accuracy_pivot, chart_row)
    wb.save(output_path)

    logger.info(f"Excel saved -> {output_path}")
    print(f"\n  Results exported to: {output_path}")
    for sname in list(data_sheets) + ["accuracy_results"]:
        print(f"   Sheet '{sname}': written")


# ============================================================================
# MAIN PIPELINE
# ============================================================================
async def run_pipeline(args: argparse.Namespace) -> int:
    print("\n" + "=" * 70)
    print("  MATRIX EVALUATION PIPELINE")
    print("=" * 70)

    os.makedirs(args.results_dir, exist_ok=True)
    os.makedirs(args.temp_dir,    exist_ok=True)

    output_xlsx     = os.path.join(args.results_dir, args.output_xlsx)
    checkpoint_path = os.path.join(args.results_dir, args.checkpoint)
    raw_backup      = os.path.join(args.temp_dir,    "master_raw")
    eval_backup     = os.path.join(args.results_dir, "master_evaluated")

    # 1. Ingest
    print("\n[1/5] Ingesting input files...")
    try:
        master_df = ingest_files(args.data_dir)
    except (FileNotFoundError, ValueError) as exc:
        logger.error(f"Ingestion failed: {exc}")
        return 1

    save_dataframe_backup(master_df, raw_backup, label="master_raw")

    # 2. Evaluate
    print("\n[2/5] Running LLM evaluations (gated by model_correct_orig)...")
    try:
        evaluated_df = await evaluate_dataframe(
            master_df,
            checkpoint_path=checkpoint_path,
            concurrency=args.concurrency,
            max_retries=args.max_retries,
            base_backoff=args.backoff,
            resume=args.resume,
            force=args.force,
            modelnum=args.model_num,
        )
    except Exception as exc:
        logger.exception(f"Evaluation failed: {exc}")
        return 1

    print("\n[3/5] Saving evaluated master backup...")
    try:
        save_dataframe_backup(evaluated_df, eval_backup, label="master_evaluated")
    except Exception as exc:
        logger.warning(f"  Backup of evaluated master failed: {exc}")

    # 3. Split
    print("\n[4/5] Splitting by format (ascii / latex / list)...")
    try:
        format_splits = split_by_format(evaluated_df)
    except Exception as exc:
        logger.exception(f"Split failed: {exc}")
        return 1

    # 4. Accuracy
    print("\n[5/5] Computing accuracy & exporting Excel...")
    try:
        accuracy_pivot = compute_accuracy(evaluated_df)
        if accuracy_pivot:
            logger.info(
                f"  Pivot: {len(accuracy_pivot['problem_types'])} problem types | "
                f"models={accuracy_pivot['models']} | formats={accuracy_pivot['formats']}"
            )
    except Exception as exc:
        logger.exception(f"Accuracy computation failed: {exc}")
        accuracy_pivot = {}

    try:
        export_to_excel(format_splits, accuracy_pivot, output_xlsx)
    except Exception as exc:
        logger.exception(f"Excel export failed: {exc}")
        return 1

    print("\n" + "=" * 70)
    print("  PIPELINE COMPLETE")
    print(f"  Excel       : {output_xlsx}")
    print(f"  Checkpoint  : {checkpoint_path}")
    print(f"  Eval backup : {eval_backup}.csv / .jsonl / .meta.json")
    print("=" * 70 + "\n")
    return 0


# ============================================================================
# CLI
# ============================================================================
def _parse_args(argv=None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Matrix Evaluation Pipeline (resumable, with JSONL backups)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=dedent("""
            Examples:
              python matrix_eval_pipeline.py
              python matrix_eval_pipeline.py --data-dir data/full --concurrency 8 --resume
              python matrix_eval_pipeline.py --model-num 2 --resume
              python matrix_eval_pipeline.py --resume --force      # ignore cache
              python matrix_eval_pipeline.py --max-retries 5 --backoff 2.0
        """).strip(),
    )
    p.add_argument("--data-dir",    default=DEFAULT_DATA_DIR,
                   help=f"Folder of input CSV/XLSX files (default: {DEFAULT_DATA_DIR})")
    p.add_argument("--temp-dir",    default=DEFAULT_TEMP_DIR,
                   help=f"Staging dir for raw backups (default: {DEFAULT_TEMP_DIR})")
    p.add_argument("--results-dir", default=DEFAULT_RESULTS_DIR,
                   help=f"Output dir for xlsx + eval backup (default: {DEFAULT_RESULTS_DIR})")
    p.add_argument("--output-xlsx", default=DEFAULT_OUTPUT_XLSX,
                   help=f"Excel filename (default: {DEFAULT_OUTPUT_XLSX})")
    p.add_argument("--checkpoint",  default=DEFAULT_CHECKPOINT,
                   help=f"JSONL checkpoint filename (default: {DEFAULT_CHECKPOINT})")
    p.add_argument("--model-num",   type=int, default=2, choices=(1, 2, 3, 4),
                   help="LLM model selector (default: 2 = openrouter/deepseek-v3.2). "
                        "1=openai gpt-5-mini, 2=deepseek-v3.2, 3=or-gpt-5-mini, 4=gpt-oss-120b")
    p.add_argument("--concurrency", type=int,   default=5,
                   help="Concurrent agent calls (default: 5)")
    p.add_argument("--max-retries", type=int,   default=3,
                   help="Retry attempts per row on transient errors (default: 3)")
    p.add_argument("--backoff",     type=float, default=1.0,
                   help="Base seconds for exponential backoff (default: 1.0)")
    p.add_argument("--resume",      action="store_true",
                   help="Skip rows already in checkpoint JSONL")
    p.add_argument("--force",       action="store_true",
                   help="Re-evaluate every row, ignoring cache and original True labels")
    p.add_argument("--debug",       action="store_true",
                   help="Verbose DEBUG logging")
    return p.parse_args(argv)


def main() -> None:
    args = _parse_args()
    if args.debug:
        setup_logging(logging.DEBUG)
    try:
        rc = asyncio.run(run_pipeline(args))
    except KeyboardInterrupt:
        logger.warning("Interrupted by user — checkpoint preserved.")
        sys.exit(130)
    sys.exit(rc)


if __name__ == "__main__":
    main()